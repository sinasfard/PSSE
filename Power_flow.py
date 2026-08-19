import sys, os, csv
# import cx_Oracle
import pandas as pd
import shutil
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import numpy as np
from collections import defaultdict
from log import MyLogger
import re
from utils_standAlone import PSSE_PATH_OS, PSSE_PATH_SYS
import zipfile
from case_standAlone import saveCase
sys.path.append(PSSE_PATH_OS)
from ShiftFactor_StandAlone import ShiftFactor
sys.path.append(PSSE_PATH_SYS)
import psspy
import math
from openpyxl import load_workbook
import Combinations
from datetime import datetime
from Voltage_Limits_Check_July2023 import ID2010_007RS
from psspy import _i, _f, _s



Albregions = {'Calgary':[57,6],
'CentralEast':[42, 36, 37, 32, 13, 56, 28],
'CentralWest':[30 ,35, 34, 38, 39],
'Edmonton':[40, 60, 31],
'NorthWest':[17, 18, 19, 20, 21, 23, 26, 24, 22],
'NorthEast':[25, 27, 33],
'SouthEast':[43, 48, 47, 52, 54, 4],
'SouthWest':[44, 45, 46, 49, 53, 55]}
OpposingRegions  = {"CentralEast":[17, 18, 19, 20,23],"CentralWest":[25,27], "Edmonton":[25,17,18] , "NorthWest":[48,4], "NorthEast":[53,46,55], "SouthEast":[25,27,18,17], "SouthWest":[25,27,18,17]}


def replace_BusnNmeWithBusNumber(row):
    if pd.isnull(row['injectionBusName']) or row['injectionBusName'] == '':
        return row['injectionBus']
    else:
        return row['injectionBusName']


def createSubFile(dirname,bus):
	template = f"""/PSS(R)E 34
				COM
				COM SUBSYSTEM description file entry created by PSS(R)E Config File Builder
				COM
				SUBSYSTEM 'TETS'
				BUS {bus}
				END
				END"""


	filename = "temp.sub"
	filename = os.path.join(dirname, filename)
	with open(filename, "w") as file:
		file.write(template)

def creatConFile(dirname):
	content = """/PSS(R)E 35
COM
COM CONTINGENCY description file entry created by PSS(R)E Config File Builder
COM
SINGLE BRANCH IN SUBSYSTEM 'TETS'
PARALLEL BRANCH IN SUBSYSTEM 'TETS'
SINGLE TIE FROM SUBSYSTEM 'TETS'
DOUBLE TIE FROM SUBSYSTEM 'TETS'
PARALLEL TIE FROM SUBSYSTEM 'TETS'
END
"""
	filename = "temp.con"
	filename = os.path.join(dirname, filename)
	with open(filename, "w") as file:
		file.write(content)


def creatMonFile(dirname):
	content = """/PSS(R)E 34
COM
COM MONITORED element file entry created by PSS(R)E Config File Builder
COM
MONITOR VOLTAGE RANGE SUBSYSTEM 'TETS' 0.950 1.050
MONITOR BRANCHES IN SUBSYSTEM 'TETS'
MONITOR TIES FROM SUBSYSTEM 'TETS'
END
"""
	filename = "temp.mon"
	filename = os.path.join(dirname, filename)
	with open(filename, "w") as file:
		file.write(content)



def createShiftFactorTable(dirname):
	rows = []
	filename = os.path.join(dirname, 'results.sf')
	with open(filename, 'r') as file:
		readFromBus = False
		readToBus = False
		readBuses = False
		fromBus = -1
		toBus = -1
		branch_id = -1
		
		for line in file:
			line_data = line.split()
			
			if "SENSITIVITY FACTORS OF BRANCH FLOW (MW) ON " in line:
				readBuses = False
				next_index = 0
				for index, str1 in enumerate(line_data):
					if next_index < len(line_data) - 1:
						next_index += 1
						if str1 == 'ON':
							readFromBus = True
							continue
						if readFromBus:
							fromBus = int(str1)
							readFromBus = False
							continue
						if str1 == "TO":
							readToBus = True
							continue
						if readToBus:
							try:
								toBus = int(str1)
							except:
								toBus = -1
							readToBus = False
							continue
						if line_data[next_index] == "MORE":
							branch_id = str1
				
			elif "<----" in line:
				readBuses = True
			elif readBuses and len(line_data) > 2:
				effectiveness = 0
				if fromBus<toBus: 
					effectiveness = round(float(line_data[-1]), 3)
				else: 
					effectiveness = -1*round(float(line_data[-1]), 3)
				rows.append({
					'injectionBus': int(line_data[0]),
					'fromBus': min(fromBus,toBus),
					'toBus': max(fromBus,toBus),
					'genEffectivness': effectiveness,
					'branchId': branch_id
				})

	return rows[0]



def createShiftFactorTableV2(dirname):
	rows = []
	filename = os.path.join(dirname, 'results.sf')
	with open(filename, 'r') as file:
		readFromBus = False
		readToBus = False
		readBuses = False
		fromBus = -1
		toBus = -1
		branch_id = -1
		
		for line in file:
			line_data = line.split()
			
			if "SENSITIVITY FACTORS OF BRANCH FLOW (MW) ON " in line:
				readBuses = False
				next_index = 0
				for index, str1 in enumerate(line_data):
					if next_index < len(line_data) - 1:
						next_index += 1
						if str1 == 'ON':
							readFromBus = True
							continue
						if readFromBus:
							fromBus = int(str1)
							readFromBus = False
							continue
						if str1 == "TO":
							readToBus = True
							continue
						if readToBus:
							try:
								toBus = int(str1)
							except:
								toBus = -1
							readToBus = False
							continue
						if line_data[next_index] == "MORE":
							branch_id = str1
				
			elif "<----" in line:
				readBuses = True
			elif readBuses and len(line_data) > 2:
				effectiveness = 0
				if fromBus<toBus: 
					effectiveness = round(float(line_data[-1]), 3)
				else: 
					effectiveness = -1*round(float(line_data[-1]), 3)
				rows.append({
					'injectionBus': int(line_data[0]),
					'fromBus': min(fromBus,toBus),
					'toBus': max(fromBus,toBus),
					'genEffectivness': effectiveness,
					'branchId': branch_id
				})
				return rows[0]



def removeFiles(dirnam):
	filename = os.path.join(dirnam, 'results.sf')
	if os.path.exists(filename):
		os.remove(filename)
	filename = os.path.join(dirnam, 'temp.mon')
	if os.path.exists(filename):
		os.remove(filename)
	filename = os.path.join(dirnam, 'temp.con')
	if os.path.exists(filename):
		os.remove(filename)
	filename = os.path.join(dirnam, 'temp.sub')
	if os.path.exists(filename):
		os.remove(filename)   
	filename = os.path.join(dirnam, 'temp.dfx')
	if os.path.exists(filename):
		os.remove(filename)   


class case_obj:
	def __init__(self, pre_case=None, curtailed_case=None):
		self.preCase = pre_case
		self.curtailedCase = curtailed_case

def find_related_cases(case_lst, case):
	string_object = case_obj()
	if case =='2028_SP_Cen-High_v3_NoPENV_CETOP1_Optimized_P2732_Stage1_alt1_solved.sav':
		print('yes')
	i = 0
	while i<len(case_lst):
		if case_lst[i].startswith(case.replace('.sav',"")) and ("Curtailed" in case_lst[i]):
			string_object.curtailedCase = case_lst[i] 
			case_lst = list(set(case_lst)-set([case_lst[i]]))
			break
		i += 1

	while len(case_lst)>=1:
		min_len_str = min(case_lst, key=len)
		min_len_str = min_len_str.replace(".sav","")
		min_len = len(min_len_str)
		string = case.replace(".sav", '')
		if string[:min_len]==min_len_str:
			if case =='2028_SP_Cen-High_v3_NoPENV_CETOP1_Optimized_P2732_Stage1_alt1_solved.sav':
				print('yes')
			string_object.preCase = min_len_str+'.sav'
			break
		case_lst = list(set(case_lst)-set([min_len_str+'.sav']))

	return string_object




def find_pre_project_case(strings, string):
	if not strings:
		return None

	while len(strings)>=1:
		min_len_str = min(strings, key=len)
		min_len_str = min_len_str.replace(".sav","")
		min_len = len(min_len_str)
		string = string.replace(".sav", '')
		if string[:min_len]==min_len_str:
			return min_len_str+'.sav'
		strings = list(set(strings)-set([min_len_str+'.sav']))

	return None  

def Sensitivity(Thermalresults, dirname, logger, areas, cons):
# def Sensitivity(results, arealist, min = 0.05, studyareaonly=True):

	ProjectGen_df = pd.read_excel(r"T:\Rshared\Delivery Shared\Python Files\GitRepo\CustomInputTemplateP2490.xlsx","ProjectInfo")
	ProjectGen_df = ProjectGen_df.loc[ProjectGen_df['AlternativeName']=='Alt0',['Project ID', 'Bus_1','MachineID1',  'Bus_2','MachineID2','Bus_3','MachineID3']]
	shiftfactors = []
	cases = Thermalresults['StudyCase'].unique().tolist()
	
	try:
		contingencies = Thermalresults['Contingency'].unique().tolist()
	except:
		contingencies = ['No Outage']
		Thermalresults['Contingency'] = 'No Outage'
	
	Thermalresults['Projects_in_Study'] = ''
	cons['No Outage'] = "N/A"

	for case in cases:
		if "WP" in case:
			pctrate = "PCTRTB"
			rate = "RATEB"
		else:
			pctrate = "PCTRTA"
			rate = "RATEA"           
		projects = re.findall(r'P\d{4}',case)  # find the list of the PXXXX from the case name 
		logger.info(projects)
		Thermalresults.loc[Thermalresults['StudyCase']==case, 'Projects_in_Study'] = '_'.join(projects)
		for contingency in contingencies:
			k = 0
			if len(projects)>0:
				for project in projects:
					GEF_col = f"{project}"
					ierr = psspy.case(os.path.join(dirname, case))
					applyCon(cons[contingency], logger)
					if Solve():
						if ierr!=0: 
							logger.warning(f"Issue with opening the {os.path.join(dirname, case)}")                
						else:
							k += 1
							if GEF_col not in list(Thermalresults.columns):
								Thermalresults[GEF_col] = None

							logger.info(f"for {case} the shift factor of follwoing project(s) are being calculated {project}")
							gen = ProjectGen_df.loc[(ProjectGen_df["Project ID"]==project),['Bus_1','MachineID1',  'Bus_2','MachineID2','Bus_3','MachineID3']]
							ierr, ival = psspy.macint(int(gen['Bus_1'].values[0]), str(gen['MachineID1'].values[0]), 'STATUS')
							if ival==1:
								gen = [int(gen['Bus_1'].values[0]), str(gen['MachineID1'].values[0])]
							elif not(gen['Bus_2'].isna().values[0]):
								ierr, ival = psspy.macint(int(gen['Bus_2'].values[0]), str(gen['MachineID2'].values[0]), 'STATUS')
								if ival==1:
									gen = [int(gen['Bus_2'].values[0]), str(gen['MachineID2'].values[0])]
								elif not(gen['Bus_3'].isna().values[0]):
									ierr, ival = psspy.macint(int(gen['Bus_3'].values[0]), str(gen['MachineID3'].values[0]), 'STATUS')
									if ival==1:
										gen = [int(gen['Bus_3'].values[0]), str(gen['MachineID3'].values[0])]                 
							if ierr==0:        
							


								psspy.bsys(0,1,[ 0, 500.],len(areas),areas,0,[],0,[],0,[])
								if len(cons)>0:
									Thermalresults_temp = Thermalresults.loc[(Thermalresults['StudyCase']==case) & (Thermalresults['Contingency']==contingency)]
								else:
									Thermalresults_temp = Thermalresults.loc[(Thermalresults['StudyCase']==case)]
								if len(Thermalresults_temp)>0:                     
									if Solve():
										Thermalresults_temp["postLoading"] = None
										Thermalresults_temp["postRating"] = None
										for i in Thermalresults_temp.index:
											try:
												createSubFile(dirname, gen[0])
												creatConFile(dirname)
												creatMonFile(dirname)  
												ierr = psspy.report_output(2,os.path.join(dirname, "results.sf"),[0,0])                                              
												psspy.bsys(0,0,[ 0.4, 500.],0,[],1,[gen[0]],0,[],0,[])
												psspy.dfax_2([1,1,0],os.path.join(dirname,"temp.sub"),
												os.path.join(dirname,"temp.mon"),
												os.path.join(dirname,"temp.con"),
												os.path.join(dirname,"temp"))
												ierr_sen = psspy.sensitivity_flow([Thermalresults_temp.loc[i]['FROMNUMBER'],Thermalresults_temp.loc[i]['TONUMBER'],0,0,0],[0,1,0,0,0,0,0,0,0],[ 0.5, 0.01],fr"""{Thermalresults_temp.loc[i]['ID']}""",[r"""TETS""",""],
												os.path.join(dirname, "temp.dfx"))                                            
												ierr = psspy.close_report()
												if ierr_sen==0:
													try:
														a = createShiftFactorTable(dirname)
														Thermalresults_temp.at[i,GEF_col]  = a['genEffectivness']*100
														if Thermalresults_temp.at[i,'MW_PostProject']<0:
															Thermalresults_temp.at[i,GEF_col] = -1*Thermalresults_temp.at[i,GEF_col]
													except:
														Thermalresults_temp.at[i,GEF_col]  = 0.0
												removeFiles(dirname)
												
											except:
												Thermalresults_temp.at[i,GEF_col] = 10000
											Thermalresults.loc[(Thermalresults['StudyCase']==case)&\
																(Thermalresults['FROMNUMBER']==Thermalresults_temp.loc[i]['FROMNUMBER'])&\
																	(Thermalresults['TONUMBER']==Thermalresults_temp.loc[i]['TONUMBER'])&\
																		(Thermalresults['ID']==Thermalresults_temp.loc[i]['ID'])&\
																		(Thermalresults['Contingency']==contingency),GEF_col]=Thermalresults_temp.at[i,GEF_col]
									else:
											ierr = 1
											logger.warning(f"""Issue with getting the GEFF of {project} for the element for bus {Thermalresults_temp.loc[i]['FROMNUMBER']} to bus {Thermalresults_temp.loc[i]['TONUMBER']}""")
											Thermalresults.loc[(Thermalresults['StudyCase']==case)&\
																	(Thermalresults['FROMNUMBER']==Thermalresults_temp.loc[i]['FROMNUMBER'])&\
																		(Thermalresults['TONUMBER']==Thermalresults_temp.loc[i]['TONUMBER'])&\
																			(Thermalresults['ID']==Thermalresults_temp.loc[i]['ID'])&\
																			(Thermalresults['Contingency']==contingency),GEF_col]=10000         
							else:
								logger.info(f"there is no shift factor to calculate for {case} and {contingency}.")                        
			else:
				logger.info(f"there is no shift factor to calculate for {case}. Please make sure the project number included in the case name.")



def Sensitivity_df(Thermalresults, dirname, logger, areas, cons, GEFF_df):
# def Sensitivity(results, arealist, min = 0.05, studyareaonly=True):
	shiftfactors = []
	cases = Thermalresults['StudyCase'].unique().tolist()
	
	try:
		contingencies = Thermalresults['Contingency'].unique().tolist()
	except:
		contingencies = ['No Outage']
		Thermalresults['Contingency'] = 'No Outage'
	
	Thermalresults['Projects_in_Study'] = ''
	cons['No Outage'] = "N/A"
	buses_dict = dict(zip(GEFF_df['Project ID'], GEFF_df['Bus Number']))
	for case in cases:
		if "WP" in case:
			pctrate = "PCTRTB"
			rate = "RATEB"
		else:
			pctrate = "PCTRTA"
			rate = "RATEA"           
		projects = re.findall(r'P\d{4}',case)  # find the list of the PXXXX from the case name 
		Thermalresults.loc[Thermalresults['StudyCase']==case, 'Projects_in_Study'] = case
		for contingency in contingencies:
			k = 0
			for project in buses_dict.keys():
				GEF_col = f'''{project}_Bus#{int(buses_dict[project])}'''
				gen_bus = int(buses_dict[project])
				ierr = psspy.case(os.path.join(dirname, case))
				applyCon(cons[contingency], logger)
				if Solve():
					if ierr!=0: 
						logger.warning(f"Issue with opening the {os.path.join(dirname, case)}")                
					else:
						k += 1
						if GEF_col not in list(Thermalresults.columns):
							Thermalresults[GEF_col] = None

						psspy.bsys(0,1,[ 0, 500.],len(areas),areas,0,[],0,[],0,[])
						if len(cons)>0:
							Thermalresults_temp = Thermalresults.loc[(Thermalresults['StudyCase']==case) & (Thermalresults['Contingency']==contingency)]
						else:
							Thermalresults_temp = Thermalresults.loc[(Thermalresults['StudyCase']==case)]
						if len(Thermalresults_temp)>0:                     
							if Solve():
								Thermalresults_temp["postLoading"] = None
								Thermalresults_temp["postRating"] = None
								for i in Thermalresults_temp.index:
									try:
										createSubFile(dirname, gen_bus)
										creatConFile(dirname)
										creatMonFile(dirname)  
										ierr = psspy.report_output(2,os.path.join(dirname, "results.sf"),[0,0])                                              
										psspy.bsys(0,0,[ 0.4, 500.],0,[],1,[gen_bus],0,[],0,[])
										psspy.dfax_2([1,1,0],os.path.join(dirname,"temp.sub"),
										os.path.join(dirname,"temp.mon"),
										os.path.join(dirname,"temp.con"),
										os.path.join(dirname,"temp"))
										ierr_sen = psspy.sensitivity_flow([Thermalresults_temp.loc[i]['FROMNUMBER'],Thermalresults_temp.loc[i]['TONUMBER'],0,0,0],[1,0,0,0,0,0,0,0,0],[ 0.5, 0.01],fr"""{Thermalresults_temp.loc[i]['ID']}""",[r"""TETS""",""],
										os.path.join(dirname, "temp.dfx"))                                            
										ierr = psspy.close_report()
										if ierr_sen==0:
											try:
												a = createShiftFactorTableV2(dirname)
												Thermalresults_temp.at[i,GEF_col]  = a['genEffectivness']*100
												if Thermalresults_temp.at[i,'MW_PostProject']<0:
													Thermalresults_temp.at[i,GEF_col] = -1*Thermalresults_temp.at[i,GEF_col]
											except:
												Thermalresults_temp.at[i,GEF_col]  = 0.0
										else: 
											logger.error(f'''Issue with calculating the effeciveness of the machine at bus {gen_bus} for branch from bus {Thermalresults_temp.loc[i]['FROMNUMBER']} to bus {Thermalresults_temp.loc[i]['TONUMBER']} and CKT {Thermalresults_temp.loc[i]['ID']}''')
											Thermalresults_temp.at[i,GEF_col] = 10000
										removeFiles(dirname)
										
									except:
										Thermalresults_temp.at[i,GEF_col] = 10000
									Thermalresults.loc[(Thermalresults['StudyCase']==case)&\
														(Thermalresults['FROMNUMBER']==Thermalresults_temp.loc[i]['FROMNUMBER'])&\
															(Thermalresults['TONUMBER']==Thermalresults_temp.loc[i]['TONUMBER'])&\
																(Thermalresults['ID']==Thermalresults_temp.loc[i]['ID'])&\
																(Thermalresults['Contingency']==contingency),GEF_col]=Thermalresults_temp.at[i,GEF_col]
							else:
									ierr = 1
									logger.warning(f"""Issue with getting the GEFF of {project} for the element for bus {Thermalresults_temp.loc[i]['FROMNUMBER']} to bus {Thermalresults_temp.loc[i]['TONUMBER']}""")
									Thermalresults.loc[(Thermalresults['StudyCase']==case)&\
															(Thermalresults['FROMNUMBER']==Thermalresults_temp.loc[i]['FROMNUMBER'])&\
																(Thermalresults['TONUMBER']==Thermalresults_temp.loc[i]['TONUMBER'])&\
																	(Thermalresults['ID']==Thermalresults_temp.loc[i]['ID'])&\
																	(Thermalresults['Contingency']==contingency),GEF_col]=10000         
					



def ReadInput(dirname):
	inputfile = os.path.join(dirname, "Input.xlsx")
	wb = load_workbook(inputfile)
	ws = wb["InputToIDEVRunner"]
	input = []
	#Proj ID, IDEV names, Common?, Alt name, T-tap, Disconnect, Area, Add?, Region, Buses
	for row in ws.values:
		tmp = []
		for c in row:
			tmp.append(str(c))
		input.append(tmp)
	return input
	
def Solve(i=0):
	islands=psspy.tree(1,-1)[1]
	while islands > 0:
		islands=psspy.tree(2,1)[1]
	psspy.fnsl([0,0,0,1,2,0,0,0])
	psspy.fnsl([0,0,0,1,2,0,0,0])
	if psspy.solved() == 0:
		return True
	elif psspy.solved()==1 and i<10:
		i += 1
		Solve(i)
		if psspy.solved()==0:
			return True
	else:
		return False        
		
def buildConList(input_dir, confile, GenerateSingleElementContingency, arealist, dirname, refCase, logger):
	cons = defaultdict(list)
	skips = []
	singlebranchtie = True
	remove_chrs = "/\\?,. "
	try:
		with open(os.path.join(input_dir,confile)) as f:
			lines = f.readlines()
		buildingCon = 0
		buildingSkip = 0
		conName = ""
		remove_chrs = "/\\?,. "
		for l in lines:
			try:
				tmp = l.split()
				if tmp != []:
					if tmp[0].lower() == "contingency":
						conName = " ".join(tmp[1:])
						
						remove_chrs = "/\\?,. " # we should remove any wiered special character in case we want to use the contingency name to save an sav file.
						for char in remove_chrs:
							conName = conName.replace(char, "")
				
						buildingCon = 1
					if tmp[0].lower() == "skip":
						buildingSkip = 1
						continue
					if buildingCon:
						if tmp[0].lower() == "open":
							if tmp[1].lower() in ("line", "branch"):
								if len(tmp)>10:
									cons[conName].append(["3wnd", int(tmp[4]), int(tmp[7]), int(tmp[10]), tmp[12].replace('"','').replace("'",'')])
								else:
									cons[conName].append(["brn", int(tmp[4]), int(tmp[7]), tmp[9].replace('"','').replace("'",'')])
							elif tmp[1].lower() == "bus":
								cons[conName].append(["disconnect", int(tmp[2])])
						elif tmp[0].lower() == "disconnect" or tmp[0].lower() == "trip"  or tmp[0].lower() == "open":
							cons[conName].append([tmp[0], int(tmp[2])])
						elif tmp[0].lower() == "remove" and tmp[1].lower() == "swshunt":
							cons[conName].append(["shunt", int(tmp[4])])
						elif tmp[0].lower() == "end":
								buildingCon = 0
						#machine contingencies are ignored from .con file as they are all added automatically
					if buildingSkip and tmp[0]!='COM':
						try:
							if len(tmp)==5:
								skips.append([int(tmp[0]), int(tmp[2]), ''.join(filter(str.isalnum,tmp[4]))])
							elif len(tmp)>5: #three windings
								skips.append([int(tmp[0]), int(tmp[2]),int(tmp[4]), ''.join(filter(str.isalnum,tmp[6]))])
						except ValueError:
							if tmp[0].lower() == "end":
								buildingSkip = 0
								
					if tmp[0].lower() == "single":
						singlebranchtie = True
			except:
				logger.warning(f"Issue with reading {l}. This line will be skipped")
	
	except:
		logger.warning("No Con file is selected or there is issue with running the confile.")
	
	if GenerateSingleElementContingency:
		
		psspy.case(os.path.join(dirname,refCase))
		print(arealist)
		err = psspy.bsys(0,1,[ 0, 500.],len(arealist),arealist,0,[],0,[],0,[])
		print("error is ", err)
		bussesmach = psspy.amachint(0, 1, ["NUMBER"])[1][0]
		idsmach = psspy.amachchar(0, 1, ["ID", "NAME"])[1]
		pmaxmach = psspy.amachreal(0, 1, ["PMAX"])[1][0]
		countmach = len(bussesmach)

		for i in range(countmach):
			bus = bussesmach[i]
			id = idsmach[0][i]
			busnam = idsmach[1][i]
			if pmaxmach[i] > 0.0:
				cons[busnam.strip() + "_" + id] = [["mach", bus, id]]
		
		
		if singlebranchtie:
			psspy.case(os.path.join(dirname,refCase))
			
			
			psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
			
			busses = psspy.abrnint(0, 1, 3, 3, 1, ["FROMNUMBER", "TONUMBER"])[1]
			ids = psspy.abrnchar(0, 1, 3, 3, 1, ["ID"])[1][0]
			countbrn = len(ids)
			
			bussestrn3 = psspy.awndint(0, 1, 3, 1, 1,  ["WIND1NUMBER", "WIND2NUMBER","WIND3NUMBER"])[1]
			idstrn3 = psspy.awndchar(0, 1, 3, 1, 1,  ["ID"])[1][0]
			counttrn3 = len(idstrn3)
			
			for i in range(countbrn):
				bus1 = busses[0][i]
				bus2 = busses[1][i]
				id = ids[i]
				if ([bus1, bus2, id] not in skips) and ([bus2, bus1, id] not in skips):
					name = "{0}".format(psspy.brnnam(bus1, bus2, id)[1])
					for char in remove_chrs:
						name = name.replace(char, "")
					if name=="":
						name = f"brn_{bus1}_{bus2}_{id}"
					cons[name] = [["brn", bus1, bus2, id]]
					
			
			for i in range (counttrn3):
				bus1 = bussestrn3[0][i]
				bus2 = bussestrn3[1][i]
				bus3 = bussestrn3[2][i]
				id = idstrn3[i]
				if ([bus1, bus2,bus3, id] not in skips) and ([bus2, bus1,bus3, id] not in skips):
					name = "{0}".format(psspy.tr3nam(bus1, bus2, bus3, id)[1])
					for char in remove_chrs:
							name = name.replace(char, "")		
					if name=="":
						name = f"brn_{bus1}_{bus2}_{bus3}_{id}"								
					cons[name] = [["3wnd", bus1, bus2, bus3, id]]
			
	
	return cons, skips
def percentVoltDeviation(V,Vmin, Vmax):
	if V<Vmin:
		return (Vmin-V)*100/Vmin
	elif V>Vmax:
		return (V-Vmax)*100/Vmax  
	else:
		return 0.0
  

def VoltageIssues( case, pre_case, arealist, dirname, cons=None):
	psspy.case(os.path.join(dirname,case))
	Solve()
	bus_numbers = psspy.abusint(-1, 1, ["NUMBER","Area"])[1]
	bus_names = psspy.abuschar(-1, 1, ["Name"])[1][0]


	bus_lookup_table = {}
	for i in range(len(bus_numbers[0])):
		key = bus_numbers[0][i]
		value = [bus_names[i].rstrip().lstrip(),bus_numbers[1][i]]
		bus_lookup_table[key] = value    
	results = []
	voltage = ID2010_007RS()
	voltage_ID2010_007RS = voltage.Voltages_dict()
	psspy.bsys(0,1,[ 400.0, 500.],len(arealist),arealist,0,[],0,[],0,[])
	busses500 = psspy.abusint(0, 1, ["NUMBER"])[1][0]
	busVoltages500 = psspy.abusreal(0, 1, ["PU"])[1][0]
	busZipped500 = zip(busses500, busVoltages500)


	psspy.bsys(1,1,[ 230., 241.],len(arealist),arealist,0,[],0,[],0,[])
	busses240 = psspy.abusint(1, 1, ["NUMBER"])[1][0]
	busVoltages240 = psspy.abusreal(1, 1, ["PU"])[1][0]
	busZipped240 = zip(busses240, busVoltages240)


	psspy.bsys(2,1,[ 143., 144.],len(arealist),arealist,0,[],0,[],0,[])
	busses144 = psspy.abusint(2, 1, ["NUMBER"])[1][0]
	busVoltages144 = psspy.abusreal(2, 1, ["PU"])[1][0]
	busZipped144 = zip(busses144, busVoltages144)




	psspy.bsys(3,1,[ 137., 138.],len(arealist),arealist,0,[],0,[],0,[])
	busses138 = psspy.abusint(3, 1, ["NUMBER"])[1][0]
	busVoltages138 = psspy.abusreal(3, 1, ["PU"])[1][0]
	busZipped138 = zip(busses138, busVoltages138)


	psspy.bsys(4,1,[ 71., 72.],len(arealist),arealist,0,[],0,[],0,[])
	busses72 = psspy.abusint(4, 1, ["NUMBER"])[1][0]
	busVoltages72 = psspy.abusreal(4, 1, ["PU"])[1][0]
	busZipped72 = zip(busses72, busVoltages72)

	
	psspy.bsys(5,1,[ 40, 69.],len(arealist),arealist,0,[],0,[],0,[])
	busses69 = psspy.abusint(5, 1, ["NUMBER"])[1][0]
	busVoltages69 = psspy.abusreal(5, 1, ["PU"])[1][0]
	busZipped69 = zip(busses69, busVoltages69)




	psspy.bsys(6,1,[ 245., 270.],len(arealist),arealist,0,[],0,[],0,[])  # this is not being used as 260kV are modeled as 240 kV in PSSE
	busses260 = psspy.abusint(6, 1, ["NUMBER"])[1][0]
	busVoltages260 = psspy.abusreal(6, 1, ["PU"])[1][0]
	busZipped260 = zip(busses260, busVoltages260)




	areaList_ATCO = list(set(arealist).intersection([26,24,17,25,19,42,20,18,22,23,21] ))
	areaList_ATCO144 = list(set(arealist).intersection([39, 32, 56, 37, 35, 28, 42, 13, 36] ))
	
	if len(areaList_ATCO)>=1:
		#ATCO 
		ATCO_buses = []
		psspy.bsys(7,1,[ 239.0, 240.],len(areaList_ATCO),areaList_ATCO,0,[],181,[7,4,23,25,21,3,1,191,80,194,58,44,42,36,10,99,59,421,13,231,11,301,192,251,321,8,381,15,5,41,73,362,371,521,29,14,
	12,391,441,40,514,190,67,69,70,79,2,30,516,9,43,416,63,312,705,82,49,33,451,39,261,110,501,472,473,193,211,72,200,888,999,523,9999,515,113,81,57,706,520,529,116,462,83,101,998,722,481,221,361,24,6,210,78,37,61,120,143,
	16,18,20,22,26,27,28,31,32,34,35,38,45,46,47,48,50,51,52,53,54,55,56,60,62,64,65,66,68,71,74,75,76,85,88,100,102,103,105,121,125,130,140,150,151,152,160,170,241,250,271,281,291,331,332,341,351,372,373,374,393,401,402,
	411,412,431,445,446,461,471,491,502,503,504,505,506,507,508,509,510,511,512,513,519],0,[])
		busses240_ATCO = psspy.abusint(7, 1, ["NUMBER"])[1][0]
		busVoltages240_ATCO = psspy.abusreal(7, 1, ["PU"])[1][0]
		busZipped240_ATCO = zip(busses240_ATCO,busVoltages240_ATCO)    
		


		psspy.bsys(8,1,[ 137.0, 138.],len(areaList_ATCO144),areaList_ATCO144,0,[],181,[7,4,23,25,21,3,1,191,80,194,58,44,42,36,10,99,59,421,13,231,11,301,192,251,321,8,381,15,5,41,73,362,371,521,29,14,
	12,391,441,40,514,190,67,69,70,79,2,30,516,9,43,416,63,312,705,82,49,33,451,39,261,110,501,472,473,193,211,72,200,888,999,523,9999,515,113,81,57,706,520,529,116,462,83,101,998,722,481,221,361,24,6,210,78,37,61,120,143,
	16,18,20,22,26,27,28,31,32,34,35,38,45,46,47,48,50,51,52,53,54,55,56,60,62,64,65,66,68,71,74,75,76,85,88,100,102,103,105,121,125,130,140,150,151,152,160,170,241,250,271,281,291,331,332,341,351,372,373,374,393,401,402,
	411,412,431,445,446,461,471,491,502,503,504,505,506,507,508,509,510,511,512,513,519],0,[])
		busses138_ATCO = psspy.abusint(8, 1, ["NUMBER"])[1][0]
		busVoltages138_ATCO = psspy.abusreal(8, 1, ["PU"])[1][0]
		busZipped138_ATCO = zip(busses138_ATCO, busVoltages138_ATCO)   
		
		
		psspy.bsys(9,1,[ 68.0, 69.],len(areaList_ATCO),areaList_ATCO,0,[],181,[7,4,23,25,21,3,1,191,80,194,58,44,42,36,10,99,59,421,13,231,11,301,192,251,321,8,381,15,5,41,73,362,371,521,29,14,
	12,391,441,40,514,190,67,69,70,79,2,30,516,9,43,416,63,312,705,82,49,33,451,39,261,110,501,472,473,193,211,72,200,888,999,523,9999,515,113,81,57,706,520,529,116,462,83,101,998,722,481,221,361,24,6,210,78,37,61,120,143,
	16,18,20,22,26,27,28,31,32,34,35,38,45,46,47,48,50,51,52,53,54,55,56,60,62,64,65,66,68,71,74,75,76,85,88,100,102,103,105,121,125,130,140,150,151,152,160,170,241,250,271,281,291,331,332,341,351,372,373,374,393,401,402,
	411,412,431,445,446,461,471,491,502,503,504,505,506,507,508,509,510,511,512,513,519],0,[])    
		busses69_ATCO = psspy.abusint(9, 1, ["NUMBER"])[1][0]
		busVoltages69_ATCO = psspy.abusreal(9, 1, ["PU"])[1][0]
		busZipped69_ATCO = zip(busses69_ATCO, busVoltages69_ATCO)
		ATCO_buses = busses69_ATCO+busses138_ATCO+busses240_ATCO
	else:
		ATCO_buses = []  
	# count = len(busses)
	# voltage_areas = []

	for entry in busZipped500:
		Vmax = 1.05
		Vmin = 1.0
		pecent = percentVoltDeviation(entry[1],Vmin, Vmax)
		if (entry[1]>Vmax or entry[1]<Vmin) and entry[0] not in ATCO_buses and entry[0] not in voltage_ID2010_007RS.keys():
			results.append([case, "No_outage", "N/A", entry[0], bus_lookup_table[entry[0]][0], bus_lookup_table[entry[0]][1], "{0:.2f} pu".format(entry[1]),"{:.1f}".format(entry[1]*500), "{:.0f}".format(Vmin*500), "{:.0f}".format(Vmax*500),"{:.0f}".format(pecent)])

	for entry in busZipped240:
		Vmax = 1.05
		Vmin = .975        
		pecent = percentVoltDeviation(entry[1],Vmin, Vmax)
		if (entry[1]>Vmax or entry[1]<Vmin)and entry[0] not in ATCO_buses and entry[0] not in voltage_ID2010_007RS.keys():
			results.append([case, "No_outage", "N/A", entry[0], bus_lookup_table[entry[0]][0], bus_lookup_table[entry[0]][1], "{0:.2f} pu".format(entry[1]),"{:.1f}".format(entry[1]*240), "{:.0f}".format(Vmin*240), "{:.0f}".format(Vmax*240),"{:.0f}".format(pecent)])

	for entry in busZipped144:
		Vmax = 1.048611
		Vmin = .95138          
		pecent = percentVoltDeviation(entry[1]*138/144,Vmin, Vmax)
		if (entry[1]>1.048611 or entry[1]<0.95138)and entry[0] not in ATCO_buses and entry[0] not in voltage_ID2010_007RS.keys():
			results.append([case, "No_outage", "N/A", entry[0], bus_lookup_table[entry[0]][0], bus_lookup_table[entry[0]][1], "{0:.2f} pu".format(entry[1]),"{:.1f}".format(entry[1]*144), "{:.0f}".format(0.95138*144), "{:.0f}".format(1.048611*144),"{:.0f}".format(pecent)])

	for entry in busZipped138:
		Vmax = 1.0507
		Vmin = .97826    	        
		pecent = percentVoltDeviation(entry[1],Vmin, Vmax)
		if (entry[1]>Vmax or entry[1]<Vmin)and entry[0] not in ATCO_buses and entry[0] not in voltage_ID2010_007RS.keys():
			results.append([case, "No_outage", "N/A", entry[0], bus_lookup_table[entry[0]][0], bus_lookup_table[entry[0]][1], "{0:.2f} pu".format(entry[1]),"{:.1f}".format(entry[1]*138), "{:.0f}".format(Vmin*138), "{:.0f}".format(Vmax*138),"{:.0f}".format(pecent)])
			
	for entry in busZipped72:
		Vmax = 1.048611
		Vmin = .95138            
		pecent = percentVoltDeviation(entry[1],Vmin, Vmax)
		if (entry[1]>Vmax or entry[1]<Vmin)and entry[0] not in ATCO_buses and entry[0] not in voltage_ID2010_007RS.keys():
			results.append([case, "No_outage", "N/A", entry[0], bus_lookup_table[entry[0]][0], bus_lookup_table[entry[0]][1], "{0:.2f} pu".format(entry[1]),"{:.1f}".format(entry[1]*72), "{:.0f}".format(Vmin*72), "{:.0f}".format(Vmax*72),"{:.0f}".format(pecent)])
			
	for entry in busZipped69:
		Vmax = 1.1
		Vmin = .949            
		pecent = percentVoltDeviation(entry[1],Vmin, Vmax)
		if (entry[1]>Vmax or entry[1]<Vmin)and entry[0] not in ATCO_buses and entry[0] not in voltage_ID2010_007RS.keys():
			results.append([case, "No_outage", "N/A", entry[0], bus_lookup_table[entry[0]][0], bus_lookup_table[entry[0]][1], "{0:.2f} pu".format(entry[1]),"{:.1f}".format(entry[1]*69), "{:.0f}".format(Vmin*69), "{:.0f}".format(Vmax*69),"{:.0f}".format(pecent)])

	for entry in busZipped260:
		Vmax = 1.023
		Vmin = .95        
		pecent = percentVoltDeviation(entry[1]*240/260 ,Vmin, Vmax)
		if (entry[1]>Vmax or entry[1]<Vmin)and entry[0] not in ATCO_buses and entry[0] not in voltage_ID2010_007RS.keys():
			results.append([case, "No_outage", "N/A", entry[0], bus_lookup_table[entry[0]][0], bus_lookup_table[entry[0]][1], "{0:.2f} pu".format(entry[1]),"{:.1f}".format(entry[1]*240), "{:.0f}".format(Vmin*260), "{:.0f}".format(Vmax*260),"{:.0f}".format(pecent)])
	if len(ATCO_buses):
		for entry in busZipped240_ATCO:
			Vmax = 1.023
			Vmin = .95             
			pecent = percentVoltDeviation(entry[1],Vmin, Vmax)
			if ((entry[1])>Vmax or (entry[1])<Vmin) and entry[0] not in voltage_ID2010_007RS.keys():
				results.append([case, "No_outage", "N/A", entry[0], bus_lookup_table[entry[0]][0], bus_lookup_table[entry[0]][1], "{0:.2f} pu".format(entry[1]),"{:.1f}".format(entry[1]*240.), "{:.0f}".format(Vmin*260), "{:.0f}".format(Vmax*260),"{:.0f}".format(pecent)])
				
		for entry in busZipped138_ATCO:
			Vmax = 1.048611
			Vmin = .95138             
			pecent = percentVoltDeviation(entry[1],Vmin, Vmax)
			if ((entry[1])>Vmax or (entry[1])<Vmin) and entry[0] not in voltage_ID2010_007RS.keys():
				results.append([case, "No_outage", "N/A", entry[0], bus_lookup_table[entry[0]][0], bus_lookup_table[entry[0]][1], "{0:.2f} pu".format(entry[1]),"{:.1f}".format(entry[1]*138.),"{:.0f}".format(Vmin*144), "{:.0f}".format(Vmax*144),"{:.0f}".format(pecent)])

		for entry in busZipped69_ATCO:
			Vmax = 1.048611
			Vmin = .95138           
			pecent = percentVoltDeviation(entry[1],Vmin, Vmax)
			if ((entry[1])>Vmax or (entry[1])<Vmin) and entry[0] not in voltage_ID2010_007RS.keys():
				results.append([case, "No_outage", "N/A", entry[0], bus_lookup_table[entry[0]][0], bus_lookup_table[entry[0]][1], "{0:.2f} pu".format(entry[1]),"{:.1f}".format(entry[1]*72.), "{:.0f}".format(Vmin*72), "{:.0f}".format(Vmax*72),"{:.0f}".format(pecent)])
	
	for bus in voltage_ID2010_007RS.keys():
		Vmax = voltage_ID2010_007RS[bus][5]
		Vmin = voltage_ID2010_007RS[bus][4]        
		ierr,PU = psspy.busdat(bus,'PU')
		ierr,kV = psspy.busdat(bus,'KV')
		ierr,area = psspy.busint(bus,'AREA')
		if area in arealist:
			if ierr!=0:
				results.append([case, "No_outage", "N/A", bus, "Bus Not Found","", "", "",""])
			
			elif  voltage_ID2010_007RS[bus][2] not in [72, 144, 260]:
				if kV < Vmin or kV > Vmax:
					pecent = percentVoltDeviation(kV,Vmin, Vmax)
					results.append([case, "No_outage", "N/A", bus, bus_lookup_table[bus][0], bus_lookup_table[bus][1], "{0:.2f} pu".format(PU),"{:.1f}".format(kV), "{:.0f}".format(Vmin), "{:.0f}".format(Vmax),"{:.0f}".format(pecent)])
			
			elif voltage_ID2010_007RS[bus][2] == 72 and (kV < Vmin or kV > Vmax):
				pecent = percentVoltDeviation(kV,Vmin, Vmax)
				results.append([case, "No_outage", "N/A", bus,  bus_lookup_table[bus][0], bus_lookup_table[bus][1],"{0:.2f} pu".format(PU),"{:.1f}".format(kV), "{:.0f}".format(Vmin), "{:.0f}".format(Vmax),"{:.0f}".format(pecent)])

			elif voltage_ID2010_007RS[bus][2] == 144 and (kV < Vmin or kV > Vmax):
				pecent = percentVoltDeviation(kV,Vmin, Vmax)
				results.append([case, "No_outage", "N/A", bus, bus_lookup_table[bus][0], bus_lookup_table[bus][1], "{0:.2f} pu".format(PU),"{:.1f}".format(kV), "{:.0f}".format(Vmin), "{:.0f}".format(Vmax),"{:.0f}".format(pecent)])  


			elif voltage_ID2010_007RS[bus][2] == 260 and ((kV) < Vmin or (kV) > Vmax):
				pecent = percentVoltDeviation(kV,Vmin, Vmax)
				results.append([case, "No_outage", "N/A", bus, bus_lookup_table[bus][0], bus_lookup_table[bus][1], "{0:.2f} pu".format(PU),"{:.1f}".format(kV), "{:.0f}".format(Vmin), "{:.0f}".format(Vmax),"{:.0f}".format(pecent)])            









			
	if cons != None:
		for con in cons:
			psspy.case(os.path.join(dirname,case))
			psspy.solution_parameters_5([_i,100,_i,_i,10,_i,_i,20,0],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
			applyCon(cons[con], logger)
			Solve()    
			if Solve():
					psspy.bsys(1,1,[ 400, 500.],len(arealist),arealist,0,[],0,[],0,[])
					busses500 = psspy.abusint(1, 1, ["NUMBER","Area"])[1]
					busVoltages500 = psspy.abusreal(1, 1, ["PU"])[1][0]
					busses_names500 = psspy.abuschar(1, 1, ["Name"])[1][0]					
					busZipped500 = zip(busses500[0],busses500[1], busVoltages500, busses_names500)


					psspy.bsys(1,1,[ 220, 400.],len(arealist),arealist,0,[],0,[],0,[])
					busses240 = psspy.abusint(1, 1, ["NUMBER","Area"])[1]
					busVoltages240 = psspy.abusreal(1, 1, ["PU"])[1][0]
					busses_names240 = psspy.abuschar(1, 1, ["Name"])[1][0]	
					busZipped240 = zip(busses240[0],busses240[1], busVoltages240, busses_names240)					
					
					psspy.bsys(1,1,[ 130.0, 220.],len(arealist),arealist,0,[],0,[],0,[])
					busses138 = psspy.abusint(1, 1, ["NUMBER","Area"])[1]
					busVoltages138 = psspy.abusreal(1, 1, ["PU"])[1][0]
					busses_names138 = psspy.abuschar(1, 1, ["Name"])[1][0]	
					busZipped138 = zip(busses138[0],busses138[1], busVoltages138, busses_names138)

					Bus_144 = [759, 1395, 1452, 1453, 1456, 1460, 1463, 1485, 1632, 1646, 549051]
					
					busses144 = psspy.bsys(1,1,[0.4,500.],len(arealist),arealist,len(Bus_144),Bus_144,0,[],0,[])
					busses144 = psspy.abusint(1, 1, ["NUMBER","Area"])[1]
					busVoltages144 = psspy.abusreal(1, 1, ["PU"])[1][0]
					busses_names144 = psspy.abuschar(1, 1, ["Name"])[1][0]	
					busZipped144 = zip(busses144[0],busses144[1], busVoltages144, busses_names144)

					
					psspy.bsys(1, 1,[ 90, 72.],len(arealist),arealist,0,[],0,[],0,[])
					busses69 = psspy.abusint(1, 1, ["NUMBER","Area"])[1]
					busVoltages69 = psspy.abusreal(1, 1, ["PU"])[1][0]
					busses_names69 = psspy.abuschar(1, 1, ["Name"])[1][0]	
					busZipped69 = zip(busses69[0],busses69[1], busVoltages69, busses_names69)
					
					voltage_areas = []
					for entry in busZipped500:
						if len(entry)>0:
							Vmax = 500*1.1
							Vmin = 500*0.95
							pecent = percentVoltDeviation(entry[2]*500,Vmin,Vmax)
							if (entry[2]>1.1 or entry[2]<0.95):
								results.append([case, con, cons[con], entry[0], entry[3], entry[1], "{0:.2f} pu".format(entry[2]), "{0:.1f}".format(entry[2]*500),Vmin,  Vmax,"{:.0f}".format(pecent)])
							
					for entry in busZipped240:
						if len(entry)>0:
							Vmax = 240*1.1
							Vmin = 240*0.9
							pecent = percentVoltDeviation(entry[2]*240,Vmin,Vmax)						
							if (entry[2]>1.1 or entry[2]<0.9):
								results.append([case, con, cons[con], entry[0], entry[3], entry[1], "{0:.2f} pu".format(entry[2]), "{0:.1f}".format(entry[2]*240),Vmin,  Vmax,"{:.0f}".format(pecent)])

					for entry in busZipped144:
						if len(entry)>0:
							if entry[0] not in busses144:
								Vmax = 138*1.12319
								Vmin = 138*0.94203
								pecent = percentVoltDeviation(entry[2]*144,Vmin,Vmax)						
								if (entry[2]>1.10145 or entry[2]<0.89855):
									results.append([case, con, cons[con], entry[0], entry[3], entry[1], "{0:.2f} pu".format(entry[2]),"{0:.1f}".format(entry[2]*138),Vmin,  Vmax,"{:.0f}".format(pecent)])		


					for entry in busZipped138:
						if len(entry)>0:
							Vmax = 138*1.10145
							Vmin = 138*.89855
							pecent = percentVoltDeviation(entry[2]*138,Vmin,Vmax)						
							if (entry[2]>1.10145 or entry[2]<0.89855):
								results.append([case, con, cons[con], entry[0], entry[3], entry[1], "{0:.2f} pu".format(entry[2]),"{0:.1f}".format(entry[2]*138), Vmin,  Vmax,"{:.0f}".format(pecent)])		


					for entry in busZipped69:
						if len(entry)>0:
							Vmax = 69*1.10145
							Vmin = 69*0.89855
							pecent = percentVoltDeviation(entry[2]*69,Vmin,Vmax)						
							if (entry[2]>1.10145 or entry[2]<0.89855):
								results.append([case, con, cons[con], entry[0], entry[3], entry[1], "{0:.2f} pu".format(entry[2]),"{0:.1f}".format(entry[2]*69), Vmin,  Vmax,"{:.0f}".format(pecent)])		
								
	
			else:
				results.append([case, con, cons[con], "Case failed to solve"])
			
	return results





def VoltageIssues_V2( case, pre_case, arealist, dirname, cons=None):
	df_merged = pd.DataFrame()
	psspy.case(os.path.join(dirname,case))
	ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
	Solve()
	bus_numbers = psspy.abusint(0, 1, ["NUMBER"])[1][0]
	bus_area = psspy.abusint(0, 1, ["AREA"])[1][0]
	bus_owner = psspy.abusint(0, 1, ["OWNER"])[1][0]
	bus_zone = psspy.abusint(0, 1, ["ZONE"])[1][0]
	bus_names = psspy.abuschar(0, 1, ["Name"])[1][0]
	bus_names = [bus_name.rstrip().lstrip() for bus_name in bus_names]
	bus_base = psspy.abusreal(0, 1, ["BASE"])[1][0]
	bus_kV = psspy.abusreal(0, 1, ["KV"])[1][0]
	bus_pu =  psspy.abusreal(0, 1, ["PU"])[1][0]
	zone_names = psspy.azonechar(-1,1,'ZONENAME')[1][0]
	zone_names = [zone_name.rstrip().lstrip() for zone_name in zone_names]
	owner_names = psspy.aownerchar(-1,1,'OWNERNAME')[1][0]
	owner_names = [owner_name.rstrip().lstrip() for owner_name in owner_names]
	zone_num = psspy.azoneint(-1,1,'NUMBER')[1][0]
	owner_num = psspy.aownerint(-1,1,'NUMBER')[1][0]
	owner = dict(zip(owner_num, owner_names))
	zone = dict(zip(zone_num, zone_names))
	bus_zone_names = [zone[zone_num] for zone_num in bus_zone]
	bus_owner_names = [owner[owner_num] for owner_num in bus_owner]
	data = {
		'Bus Number': bus_numbers,
		'Area': bus_area,
		'Owner Num': bus_owner,
		'Owner Name': bus_owner_names,
		'Zone': bus_zone,
		'Zone Name': bus_zone_names,
		'Name': bus_names,
		'Base Voltage (kV)': bus_base,
		'Voltage (kV)-post': bus_kV,
		'Voltage (PU)-post': bus_pu
	}
	project_df = pd.DataFrame(data)	
	pre_project_df = pd.DataFrame()
	if (not (pre_case is None)) and pre_case!=case:

	
		psspy.case(os.path.join(dirname,pre_case))
		ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
		Solve()
		bus_numbers = psspy.abusint(0, 1, ["NUMBER"])[1][0]
		bus_area = psspy.abusint(0, 1, ["AREA"])[1][0]
		bus_owner = psspy.abusint(0, 1, ["OWNER"])[1][0]
		bus_zone = psspy.abusint(0, 1, ["ZONE"])[1][0]
		bus_names = psspy.abuschar(0, 1, ["Name"])[1][0]
		bus_names = [bus_name.rstrip().lstrip() for bus_name in bus_names]
		bus_base = psspy.abusreal(0, 1, ["BASE"])[1][0]
		bus_kV = psspy.abusreal(0, 1, ["KV"])[1][0]
		bus_pu =  psspy.abusreal(0, 1, ["PU"])[1][0]

	
		data = {
			'Bus Number': bus_numbers,
			'Base Voltage (kV)': bus_base,
			'Voltage (kV)-pre': bus_kV,
			'Voltage (PU)-pre': bus_pu
		}
		pre_project_df = pd.DataFrame(data)
	if len(pre_project_df)>0:
		df_merged = pd.merge(project_df, pre_project_df, how='left', on=['Bus Number','Base Voltage (kV)'])
	else:
		df_merged = project_df
		df_merged = pd.concat([df_merged,pd.DataFrame(columns=list(['Voltage (kV)-pre','Voltage (PU)-pre']))])

	df_merged['Contingency']="Base Case"
	df_merged['Post Project Case']= case 
	df_merged['Pre Project Case']= pre_case 
	
			
	if cons != None:
		for con in cons:
			psspy.case(os.path.join(dirname,case))
			psspy.solution_parameters_5([_i,100,_i,_i,10,_i,_i,20,0],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
			applyCon(cons[con], logger)   
			ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
			if Solve():
				bus_numbers = psspy.abusint(0, 1, ["NUMBER"])[1][0]
				bus_area = psspy.abusint(0, 1, ["AREA"])[1][0]
				bus_owner = psspy.abusint(0, 1, ["OWNER"])[1][0]
				bus_zone = psspy.abusint(0, 1, ["ZONE"])[1][0]
				bus_names = psspy.abuschar(0, 1, ["Name"])[1][0]
				bus_names = [bus_name.rstrip().lstrip() for bus_name in bus_names]
				bus_base = psspy.abusreal(0, 1, ["BASE"])[1][0]
				bus_kV = psspy.abusreal(0, 1, ["KV"])[1][0]
				bus_pu =  psspy.abusreal(0, 1, ["PU"])[1][0]

				
				data = {
					'Bus Number': bus_numbers,
					'Base Voltage (kV)': bus_base,
					'Voltage (kV)-post': bus_kV,
					'Voltage (PU)-post': bus_pu,
					'Contingency':con
				}
				project_con_df = pd.DataFrame(data)	
				pre_project_con_df = pd.DataFrame()
				if not ((pre_case is None) and pre_case!=case):

				
					psspy.case(os.path.join(dirname,pre_case))
					ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
					if Solve():
						bus_numbers = psspy.abusint(0, 1, ["NUMBER"])[1][0]
						bus_area = psspy.abusint(0, 1, ["AREA"])[1][0]
						bus_owner = psspy.abusint(0, 1, ["OWNER"])[1][0]
						bus_zone = psspy.abusint(0, 1, ["ZONE"])[1][0]
						bus_names = psspy.abuschar(0, 1, ["Name"])[1][0]
						bus_names = [bus_name.rstrip().lstrip() for bus_name in bus_names]
						bus_base = psspy.abusreal(0, 1, ["BASE"])[1][0]
						bus_kV = psspy.abusreal(0, 1, ["KV"])[1][0]
						bus_pu =  psspy.abusreal(0, 1, ["PU"])[1][0]

					
						data = {
							'Bus Number': bus_numbers,
							'Base Voltage (kV)': bus_base,
							'Voltage (kV)-pre': bus_kV,
							'Voltage (PU)-pre': bus_pu
						}
						pre_project_con_df = pd.DataFrame(data)
						df_con_merged = pd.merge(project_con_df, pre_project_con_df, how='left', on=['Bus Number','Base Voltage (kV)'])
				else:
					df_con_merged = pd.concat([project_con_df,pd.DataFrame(columns=list(['Voltage (kV)-pre','Voltage (PU)-pre']))])

				df_merged = pd.concat([df_merged, df_con_merged], ignore_index=True)
			else:
				data = {
					'Bus Number': "Case failed to solve",
					'Area': "Case failed to solve",
					'Owner Num': "Case failed to solve",
					'Owner Name': "Case failed to solve",
					'Zone': "Case failed to solve",
					'Zone Name': bus_zone_names,
					'Name': "Case failed to solve",
					'Base Voltage (kV)': "Case failed to solve",
					'Voltage (kV)-pre': "Case failed to solve",
					'Voltage (PU)-pre': "Case failed to solve"
				}		

				df_merged.append([case, con, cons[con], "Case failed to solve"])


	results = []
	voltage = ID2010_007RS()
	voltage_ID2010_007RS = voltage.Voltages_dict()
	voltage_ID2010_007RS_df = list(voltage_ID2010_007RS.values())
	columns = ['Name', 'Bus Number', 'Base Voltage (kV)', 'MinEmergency', 'MinNormal', 'MaxNormal', 'MaxEmergency']
	voltage_ID2010_007RS_df = pd.DataFrame(voltage_ID2010_007RS_df, columns=columns)

	voltage_ID2010_007RS_df = voltage_ID2010_007RS_df.drop(columns=['Name','Base Voltage (kV)'])
	df_merged = pd.merge(df_merged, voltage_ID2010_007RS_df, how='left', on=['Bus Number'])

	# df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 33, 60, 27, 25,20, 22, 19, 17, 18, 21, 23]))*(df_merged['Owner Num'].isin([1, 4, 25, 65, 371, 391]))*(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MinEmergency']] = 130.0
	# df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 33, 60, 27, 25,20, 22, 19, 17, 18, 21, 23]))*(df_merged['Owner Num'].isin([1, 4, 25, 65, 371, 391]))*(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MinNormal']] = 137.0
	# df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 33, 60, 27, 25,20, 22, 19, 17, 18, 21, 23]))*(df_merged['Owner Num'].isin([1, 4, 25, 65, 371, 391]))*(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MaxNormal']] = 151.0
	# df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 33, 60, 27, 25,20, 22, 19, 17, 18, 21, 23]))*(df_merged['Owner Num'].isin([1, 4, 25, 65, 371, 391]))*(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MaxEmergency']] = 155.0


	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 33, 60, 27, 25,20, 22, 19, 17, 18, 21, 23]))*(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MinEmergency']] = 130.0
	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 33, 60, 27, 25,20, 22, 19, 17, 18, 21, 23]))*(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MinNormal']] = 137.0
	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 33, 60, 27, 25,20, 22, 19, 17, 18, 21, 23]))*(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MaxNormal']] = 151.0
	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 33, 60, 27, 25,20, 22, 19, 17, 18, 21, 23]))*(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MaxEmergency']] = 155.0



	# df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 25, 28, 42, 13, 36, 33, 60, 27, 25]))*(df_merged['Owner Num'].isin([1, 4, 25, 65, 371, 391]))*(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MinEmergency']] = 234.0
	# df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 25, 42, 13, 36, 33, 60, 27, 25]))*(df_merged['Owner Num'].isin([1, 4, 25, 65, 371, 391]))*(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MinNormal']] = 247.0
	# df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 25, 13, 36, 33, 60, 27, 25]))*(df_merged['Owner Num'].isin([1, 4, 25, 65, 371, 391]))*(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MaxNormal']] = 266.0
	# df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 25, 13, 36, 33, 60, 27, 25]))*(df_merged['Owner Num'].isin([1, 4, 25, 65, 371, 391]))*(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MaxEmergency']] = 275.0



	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 25, 28, 42, 13, 36, 33, 60, 27]))*(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MinEmergency']] = 234.0
	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 25, 42, 13, 36, 33, 60, 27]))*(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MinNormal']] = 247.0
	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 25, 13, 36, 33, 60, 27]))*(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MaxNormal']] = 266.0
	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 25, 13, 36, 33, 60, 27]))*(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MaxEmergency']] = 275.0



	# df_merged.loc[(df_merged['Area'].isin([60, 21, 26, 25]))*(df_merged['Owner Num'].isin([1, 5, 99]))*(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MinEmergency']] = 65.0
	# df_merged.loc[(df_merged['Area'].isin([60, 21, 26, 25]))*(df_merged['Owner Num'].isin([1, 5, 99]))*(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MinNormal']] = 68.5
	# df_merged.loc[(df_merged['Area'].isin([60, 21, 26, 25]))*(df_merged['Owner Num'].isin([1, 5, 99]))*(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MaxNormal']] = 75.5
	# df_merged.loc[(df_merged['Area'].isin([60, 21, 26, 25]))*(df_merged['Owner Num'].isin([1, 5, 99]))*(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MaxEmergency']] = 79.0



	df_merged.loc[(df_merged['Area'].isin([60, 21, 26, 25]))*(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MinEmergency']] = 65.0
	df_merged.loc[(df_merged['Area'].isin([60, 21, 26, 25]))*(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MinNormal']] = 68.5
	df_merged.loc[(df_merged['Area'].isin([60, 21, 26, 25]))*(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MaxNormal']] = 75.5
	df_merged.loc[(df_merged['Area'].isin([60, 21, 26, 25]))*(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MaxEmergency']] = 79.0



	# 
	df_merged.loc[(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MinEmergency']] = 216.0
	df_merged.loc[(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MinNormal']] = 234.0
	df_merged.loc[(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MaxNormal']] = 252.0
	df_merged.loc[(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MaxEmergency']] = 264.0

	df_merged.loc[(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MinEmergency']] = 124.0
	df_merged.loc[(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MinNormal']] = 135.0
	df_merged.loc[(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MaxNormal']] = 145.0
	df_merged.loc[(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MaxEmergency']] = 152.0

	df_merged.loc[(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MinEmergency']] = 62.0
	df_merged.loc[(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MinNormal']] = 65.5
	df_merged.loc[(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MaxNormal']] = 72.5
	df_merged.loc[(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MaxEmergency']] = 76
	
	df_merged = df_merged[df_merged['Base Voltage (kV)'] >= 69.0]
	
	df_merged = df_merged.loc[(df_merged['Voltage (kV)-post']<df_merged['MinNormal']) | (df_merged['Voltage (kV)-post']>df_merged['MaxNormal'])|(df_merged['Voltage (kV)-pre']<df_merged['MinNormal']) | (df_merged['Voltage (kV)-pre']>df_merged['MaxNormal'])]

	return df_merged





def VoltageValues(arealist, isPre=False, con_name=None):
	ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
	bus_numbers = psspy.abusint(0, 1, ["NUMBER"])[1][0]
	bus_area = psspy.abusint(0, 1, ["AREA"])[1][0]
	bus_owner = psspy.abusint(0, 1, ["OWNER"])[1][0]
	bus_zone = psspy.abusint(0, 1, ["ZONE"])[1][0]
	bus_names = psspy.abuschar(0, 1, ["Name"])[1][0]
	bus_names = [bus_name.rstrip().lstrip() for bus_name in bus_names]
	bus_base = psspy.abusreal(0, 1, ["BASE"])[1][0]
	bus_kV = psspy.abusreal(0, 1, ["KV"])[1][0]
	bus_pu =  psspy.abusreal(0, 1, ["PU"])[1][0]
	zone_names = psspy.azonechar(-1,1,'ZONENAME')[1][0]
	zone_names = [zone_name.rstrip().lstrip() for zone_name in zone_names]
	owner_names = psspy.aownerchar(-1,1,'OWNERNAME')[1][0]
	owner_names = [owner_name.rstrip().lstrip() for owner_name in owner_names]
	zone_num = psspy.azoneint(-1,1,'NUMBER')[1][0]
	owner_num = psspy.aownerint(-1,1,'NUMBER')[1][0]
	owner = dict(zip(owner_num, owner_names))
	zone = dict(zip(zone_num, zone_names))
	bus_zone_names = [zone[zone_num] for zone_num in bus_zone]
	bus_owner_names = [owner[owner_num] for owner_num in bus_owner]
	if isPre:
		data = {
			'Bus Number': bus_numbers,
			'Base Voltage (kV)': bus_base,
			'Voltage (kV)-pre': bus_kV,
			'Voltage (PU)-pre': bus_pu,
		}		
	else:
		data = {
			'Bus Number': bus_numbers,
			'Area': bus_area,
			'Owner Num': bus_owner,
			'Owner Name': bus_owner_names,
			'Zone': bus_zone,
			'Zone Name': bus_zone_names,
			'Name': bus_names,
			'Base Voltage (kV)': bus_base,
			'Voltage (kV)-post': bus_kV,
			'Voltage (PU)-post': bus_pu,
			'Contingency':con_name
		}
	
	project_df = pd.DataFrame(data)	
	
	return project_df

def VoltageCehck(pre_df,post_df,cons=None):
	
	df_merged = pd.merge(post_df, pre_df, how='left', on=['Bus Number','Base Voltage (kV)'])
	voltage = ID2010_007RS()
	voltage_ID2010_007RS = voltage.Voltages_dict()
	voltage_ID2010_007RS_df = list(voltage_ID2010_007RS.values())
	columns = ['Name', 'Bus Number', 'Base Voltage (kV)', 'MinEmergency', 'MinNormal', 'MaxNormal', 'MaxEmergency']
	voltage_ID2010_007RS_df = pd.DataFrame(voltage_ID2010_007RS_df, columns=columns)

	voltage_ID2010_007RS_df = voltage_ID2010_007RS_df.drop(columns=['Name','Base Voltage (kV)'])
	df_merged = pd.merge(df_merged, voltage_ID2010_007RS_df, how='left', on=['Bus Number'])

	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 60, 27, 25,20, 22, 19, 17, 18, 21, 23]))*(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MinEmergency']] = 130.0
	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 60, 27, 25,20, 22, 19, 17, 18, 21, 23]))*(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MinNormal']] = 137.0
	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 60, 27, 25,20, 22, 19, 17, 18, 21, 23]))*(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MaxNormal']] = 151.0
	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 60, 27, 25,20, 22, 19, 17, 18, 21, 23]))*(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MaxEmergency']] = 155.0

	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 60, 27, 40, 25]))*(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MinEmergency']] = 234.0
	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 27, 25]))*(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MinNormal']] = 247.0
	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 27, 25]))*(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MaxNormal']] = 266.0
	df_merged.loc[(df_merged['Area'].isin([39, 32, 56, 37, 35, 28, 42, 13, 36, 60, 27, 40, 25]))*(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MaxEmergency']] = 275.0


	# df_merged.loc[(df_merged['Area'].isin([40]))*(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MinEmergency']] = 230.0
	df_merged.loc[(df_merged['Area'].isin([40]))*(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MinNormal']] = 239.0
	df_merged.loc[(df_merged['Area'].isin([40]))*(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MaxNormal']] = 260.0
	df_merged.loc[(df_merged['Area'].isin([60]))*(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MinNormal']] = 238.0
	df_merged.loc[(df_merged['Area'].isin([60]))*(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MaxNormal']] = 255.0
	# df_merged.loc[(df_merged['Area'].isin([40]))*(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MaxEmergency']] = 260.0




	df_merged.loc[(df_merged['Area'].isin([60, 21, 26, 25]))*(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MinEmergency']] = 65.0
	df_merged.loc[(df_merged['Area'].isin([60, 21, 26, 25]))*(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MinNormal']] = 68.5
	df_merged.loc[(df_merged['Area'].isin([60, 21, 26, 25]))*(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MaxNormal']] = 75.5
	df_merged.loc[(df_merged['Area'].isin([60, 21, 26, 25]))*(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MaxEmergency']] = 79.0



	# 
	df_merged.loc[(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MinEmergency']] = 216.0
	df_merged.loc[(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MinNormal']] = 234.0
	df_merged.loc[(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MaxNormal']] = 252.0
	df_merged.loc[(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==240.0), ['MaxEmergency']] = 264.0

	df_merged.loc[(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MinEmergency']] = 124.0
	df_merged.loc[(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MinNormal']] = 135.0
	df_merged.loc[(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MaxNormal']] = 145.0
	df_merged.loc[(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==138.0), ['MaxEmergency']] = 152.0

	df_merged.loc[(df_merged['MinEmergency'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MinEmergency']] = 62.0
	df_merged.loc[(df_merged['MinNormal'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MinNormal']] = 65.5
	df_merged.loc[(df_merged['MaxNormal'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MaxNormal']] = 72.5
	df_merged.loc[(df_merged['MaxEmergency'].isna())*(df_merged['Base Voltage (kV)']==69.0), ['MaxEmergency']] = 76
	
	df_merged = df_merged[df_merged['Base Voltage (kV)'] >= 69.0]
	if cons==None:
		df_merged = df_merged.loc[(df_merged['Voltage (kV)-post']<df_merged['MinNormal']) | (df_merged['Voltage (kV)-post']>df_merged['MaxNormal'])|(df_merged['Voltage (kV)-pre']<df_merged['MinNormal']) | (df_merged['Voltage (kV)-pre']>df_merged['MaxNormal'])]
	else:
		df_merged = df_merged.loc[(df_merged['Voltage (kV)-post']<df_merged['MinEmergency']) | (df_merged['Voltage (kV)-pre']>df_merged['MaxEmergency'])]
	return df_merged







def applyCon(con, logger):
	for conEntry in con:
		if conEntry[0].lower() == "disconnect" or conEntry[0].lower() == "trip" or conEntry[0].lower() == "open":
			ierr = psspy.dscn(conEntry[1])
			if ierr==0:
				logger.info(f"Bus {conEntry[1]} is tripped.")
			else:
				logger.error(f"Issue with tripping bus {conEntry[1]}.")
		elif conEntry[0].lower() == "brn":
			
			ierr1 = psspy.branch_chng_3(conEntry[1],conEntry[2],conEntry[3],[0,_i,_i,_i,_i,_i],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],[ _f, _f, _f, _f,_f,_f,_f,_f,_f,_f,_f,_f],_s)
			if ierr1==0:
				logger.info(f"Branch {conEntry[1]}_{conEntry[2]}_{conEntry[3]} is disconnected.")
			else:
				ierr2, realaro = psspy.two_winding_chng_6(conEntry[1],conEntry[2],conEntry[3],[0,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],[ _f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],_s,_s)
				if ierr2==0:
					logger.info(f"Transformer {conEntry[1]}_{conEntry[2]}_{conEntry[3]} is disconnected.")
				else:
					logger.error("Issue with tripping branch or 2 wind transformer")  
					           
		elif conEntry[0].lower() == "3wnd":
			ierr, realaro = psspy.three_wnd_imped_chng_4(conEntry[1],conEntry[2],conEntry[3],conEntry[4],[_i,_i,_i,_i,_i,_i,_i,0,_i,_i,_i,_i,_i],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],_s,_s)

			if ierr==0:
				logger.info(f"Three winding Transformer between buses: {conEntry[1]}, {conEntry[2]}, and {conEntry[3]} is disconnected.")
			else:
				logger.error(f"Issue with triping three winding Transformer between buses: {conEntry[1]}, {conEntry[2]}, and {conEntry[3]}.")      

		elif conEntry[0].lower() == "mach":
			print( conEntry)
			ierr= psspy.machine_chng_2(conEntry[1],conEntry[2],[0,_i,_i,_i,_i,_i],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
			if ierr==0:
				logger.info(f"Machine at bus {conEntry[1]} with ID {conEntry[2]} is disconnected.")
			else:
				logger.error(f"Issue with triping Machine at bus {conEntry[1]} with ID {conEntry[2]}.")  

		elif conEntry[0].lower()=="shunt":
			ierr = psspy.switched_shunt_chng_4(conEntry[1],[_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,0,0,_i],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],_s)
			if ierr==0:
				logger.info(f"Switched shunt at bus {conEntry[1]} is disconnected.")
			else: logger.error(f"Issue with disconnecting switched shunt at bus {conEntry[1]}.")  
		else:
			logger.error("Unknown command.") 



def BranchOverloads_v3(dirname, case, pre_case,  postCurtailment, arealist, cons, Limit_Check, logger = None): # this is a modified version of BranchOverloads which can populate df for Category A violations in which pre and post are compared against each other
	
	if logger is not None: logger.info(f"Running power flow for {case}, {pre_case},{postCurtailment}")
	psspy.case(os.path.join(dirname, case))
	Solve()
	if not Solve():
		if not logger is None: logger.error("The case is not solved!")
	ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
	if 'WP' in case:
		flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEB", "RATEB", "RATE3","P"])[1]
	else:
		flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEA", "RATEA", "RATE3","P"])[1]
	busses = psspy.abrnint(0, 1, 3, 3, 1, ["FROMNUMBER", "TONUMBER"])[1]
	ids = psspy.abrnchar(0, 1, 3, 3, 1, ["ID", "FROMNAME", "TONAME"])[1]
	names = [psspy.brnnam(busses[0][i], busses[1][i], f"{ids[0][i]}")[1].replace(" ", "") for i in range(len(ids[0]))]
	actualFlows = [round(flows[0][i]*flows[1][i]/100.0,2) for i in range(len(ids[0]))]
	if cons==None:
		data = {'FROMNAME': ids[1][:],'TONAME': ids[2][:], 'ID': ids[0][:], 'NAME':names, 'NORMALRATING':flows[1],'POWERFLOW':actualFlows,'LOADING': flows[0], 'FROMNUMBER':busses[0], 'TONUMBER':busses[1], "MW":flows[3]}
	else:
		data = {'FROMNAME': ids[1][:],'TONAME': ids[2][:], 'ID': ids[0][:], 'NAME':names, 'NORMALRATING':flows[1],'EmergencyRating':flows[2],'POWERFLOW':actualFlows,'LOADING': flows[0], 'FROMNUMBER':busses[0], 'TONUMBER':busses[1], "MW":flows[3]}
	
	PostProject_df = pd.DataFrame(data)
	PostProject_Volt_df = VoltageValues(arealist, False, 'No Outage')

	


	PreProject_df = pd.DataFrame(columns=['ID','FROMNUMBER','TONUMBER','LOADING','POWERFLOW', 'MW'])
	PreProject_Volt_df = pd.DataFrame(columns=[	'Bus Number',	'Base Voltage (kV)',	'Voltage (kV)-pre',	'Voltage (PU)-pre'])
	if not pre_case is None:    
		psspy.case(os.path.join(dirname, pre_case))
		Solve()
		if not Solve():
			if not logger is None: 
				logger.error(f"The case is not solved for {pre_case}!")        
		ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
		if 'WP' in case:
			flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEB", "RATEB", "RATE3","P"])[1]
		else:
			flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEA", "RATEA", "RATE3","P"])[1]


		busses = psspy.abrnint(0, 1, 3, 3, 1, ["FROMNUMBER", "TONUMBER"])[1]
		ids = psspy.abrnchar(0, 1, 3, 3, 1, ["ID", "FROMNAME", "TONAME"])[1]
		names = [psspy.brnnam(busses[0][i], busses[1][i], f"{ids[0][i]}")[1].replace(" ", "") for i in range(len(ids[0]))]
		actualFlows = [round(flows[0][i]*flows[1][i]/100.0,2) for i in range(len(ids[0]))]
		data = {'ID': ids[0][:],'POWERFLOW':actualFlows,'LOADING': flows[0], 'FROMNUMBER':busses[0], 'TONUMBER':busses[1], 'MW':flows[3], 'FROMNUMBER_Pre':busses[0], 'TONUMBER_Pre':busses[1]}
		PreProject_df = pd.DataFrame(data)
		PreProject_Volt_df = VoltageValues(arealist, True, 'No Outage')
	volt_df = VoltageCehck(PreProject_Volt_df, PostProject_Volt_df, cons)
	merged_volt_df = volt_df
	merged_volt_df['StudyCase'] = case





	PostCurtailmentProject_df = pd.DataFrame(columns=['ID','FROMNUMBER','TONUMBER','LOADING','POWERFLOW'])
	if not postCurtailment is None:    
		psspy.case(os.path.join(dirname, postCurtailment))
		Solve()
		if not Solve():
			if not logger is None: logger.error("The case is not solved!")        
		ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
		if 'WP' in case:
			flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEB", "RATEB", "RATE3","P"])[1]
		else:
			flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEA", "RATEA", "RATE3","P"])[1]


		busses = psspy.abrnint(0, 1, 3, 3, 1, ["FROMNUMBER", "TONUMBER", "STATUS"])[1]
		ids = psspy.abrnchar(0, 1, 3, 3, 1, ["ID", "FROMNAME", "TONAME"])[1]
		names = [psspy.brnnam(busses[0][i], busses[1][i], f"{ids[0][i]}")[1].replace(" ", "") for i in range(len(ids[0]))]
		actualFlows = [round(flows[0][i]*flows[1][i]/100.0,2) for i in range(len(ids[0]))]
		data = {'ID': ids[0][:],'POWERFLOW':actualFlows,'LOADING': flows[0], 'FROMNUMBER':busses[0], 'TONUMBER':busses[1], 'MW':flows[3], 'FROMNUMBER_curtail':busses[0], 'TONUMBER_curtail':busses[1]}
		PostCurtailmentProject_df = pd.DataFrame(data)        


	# PostProject_df = PostProject_df.loc[PostProject_df['LOADING']>99.9]
	# PreProject_df = PreProject_df.loc[PreProject_df['LOADING']>99.9] 
	merged_df = pd.merge(PostProject_df, PreProject_df, on=['ID','FROMNUMBER','TONUMBER'], how='left', suffixes=('_x', '_y'))
	merged_df = merged_df.rename(columns={'LOADING_x': 'LOADING_PostProject', 'LOADING_y': 'LOADING_PreProject', 'POWERFLOW_x': 'POWERFLOW_PostProject', 'POWERFLOW_y': 'POWERFLOW_PreProject', 'MW_x': 'MW_PostProject','MW_y': 'MW_PreProject' })
	
	merged_df = pd.merge(merged_df, PostCurtailmentProject_df, on=['ID','FROMNUMBER','TONUMBER'], how='left', suffixes=('_x', '_y'))
	
	# merged_df = pd.merge(merged_df, PostCurtailmentProject_df, on=['ID'], how='left', suffixes=('_x', '_y'))
	merged_df = merged_df.rename(columns={'LOADING': 'LOADING_PostCurtailment','POWERFLOW':'POWERFLOW_PostCurtailment', 'MW' : 'MW_PostCurtailment'})
	# tem_to_test = os.path.join(dirname, merged_df)
	# df.to_csv(tem_to_test)
	results = merged_df.loc[(merged_df['LOADING_PostProject']>Limit_Check) | (merged_df['LOADING_PreProject']>Limit_Check)| (merged_df['LOADING_PostCurtailment']>Limit_Check)]
	
	
		


	psspy.case(os.path.join(dirname, case))
	Solve()
	if not Solve():
		if not logger is None: logger.error("The case is not solved!")        

	if 'WP' in case:
		flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEB", "RATEB", "RATE3","P"])[1]
	else:
		flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEA", "RATEA", "RATE3","P"])[1]
	bussestrn3 = psspy.awndint(0, 1, 3, 1, 1,  ["WIND1NUMBER", "WIND2NUMBER","WIND3NUMBER", "WNDNUM"])[1]
	idstrn3 = psspy.awndchar(0, 1, 3, 1, 1,  ["ID","WIND1NAME","WIND2NAME","WIND3NAME"])[1]
	names = [psspy.tr3nam(bussestrn3[0][i], bussestrn3[1][i],bussestrn3[2][i], f"{idstrn3[0][i]}")[1].replace(" ", "") for i in range(len(idstrn3[0]))]
	actualFlows = [round(flowstrn3[0][i]*flowstrn3[1][i]/100.0,2) for i in range(len(idstrn3[0]))]
	if cons==None:
		data = {'FROMNAME': idstrn3[1][:],'TONAME': idstrn3[2][:], 'TONAME2': idstrn3[3][:], 'ID': idstrn3[0][:], 'NAME':names, 'NORMALRATING':flowstrn3[1],'POWERFLOW':actualFlows,'LOADING': flowstrn3[0], 'FROMNUMBER':bussestrn3[0], 'TONUMBER':bussestrn3[1], 'TONUMBER2':bussestrn3[2], "MW":"N/A"}
	else:
		data = {'FROMNAME': idstrn3[1][:],'TONAME': idstrn3[2][:], 'TONAME2': idstrn3[3][:], 'ID': idstrn3[0][:], 'NAME':names, 'NORMALRATING':flowstrn3[1], 'EmergencyRating': flowstrn3[2], 'POWERFLOW':actualFlows,'LOADING': flowstrn3[0], 'FROMNUMBER':bussestrn3[0], 'TONUMBER':bussestrn3[1], 'TONUMBER2':bussestrn3[2], "MW":"N/A"}
	PostProject_df = pd.DataFrame(data)

	PreProject_df = pd.DataFrame(columns=['ID','FROMNUMBER','TONUMBER','TONUMBER2','LOADING','POWERFLOW', "MW"])
	if not pre_case is None:    
		psspy.case(os.path.join(dirname, pre_case))
		Solve()
		if not Solve():
			if not logger is None: logger.error(f"The case is not solved for {pre_case}.")            
		ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
		if 'WP' in case:
			flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEB", "RATEB"])[1]
		else:
			flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEA", "RATEA"])[1]
		bussestrn3 = psspy.awndint(0, 1, 3, 1, 1,  ["WIND1NUMBER", "WIND2NUMBER","WIND3NUMBER", "WNDNUM"])[1]
		idstrn3 = psspy.awndchar(0, 1, 3, 1, 1,  ["ID","WIND1NAME","WIND2NAME","WIND3NAME"])[1]
		names = [psspy.tr3nam(bussestrn3[0][i], bussestrn3[1][i],bussestrn3[2][i], f"{idstrn3[0][i]}")[1].replace(" ", "") for i in range(len(idstrn3[0]))]
		actualFlows = [round(flowstrn3[0][i]*flowstrn3[1][i]/100.0,2) for i in range(len(idstrn3[0]))]
		data = {'ID': idstrn3[0][:], 'POWERFLOW':actualFlows,'LOADING': flowstrn3[0], 'FROMNUMBER':bussestrn3[0], 'TONUMBER':bussestrn3[1], 'TONUMBER2':bussestrn3[2], "MW":"N/A", 'FROMNUMBER_Pre':bussestrn3[0], 'TONUMBER_Pre':bussestrn3[1], 'TONUMBER2_Pre':bussestrn3[2]}
		PreProject_df = pd.DataFrame(data)


	PostCurtailmentProject_df = pd.DataFrame(columns=['ID','FROMNUMBER','TONUMBER','TONUMBER2','LOADING','POWERFLOW',"MW"])
	if not postCurtailment is None:    
		psspy.case(os.path.join(dirname, postCurtailment))
		Solve()
		if not Solve():
			if not logger is None: logger.error("The case is not solved!")            
		ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
		if 'WP' in case:
			flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEB", "RATEB"])[1]
		else:
			flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEA", "RATEA"])[1]
		bussestrn3 = psspy.awndint(0, 1, 3, 1, 1,  ["WIND1NUMBER", "WIND2NUMBER","WIND3NUMBER", "WNDNUM", "STATUS"])[1]
		idstrn3 = psspy.awndchar(0, 1, 3, 1, 1,  ["ID","WIND1NAME","WIND2NAME","WIND3NAME"])[1]
		names = [psspy.tr3nam(bussestrn3[0][i], bussestrn3[1][i],bussestrn3[2][i], f"{idstrn3[0][i]}")[1].replace(" ", "") for i in range(len(idstrn3[0]))]
		actualFlows = [round(flowstrn3[0][i]*flowstrn3[1][i]/100.0,2) for i in range(len(idstrn3[0]))]
		data = {'ID': idstrn3[0][:], 'POWERFLOW':actualFlows,'LOADING': flowstrn3[0], 'FROMNUMBER':bussestrn3[0], 'TONUMBER':bussestrn3[1], 'TONUMBER2':bussestrn3[2], "MW":"N/A", 'FROMNUMBER_curtail':bussestrn3[0], 'TONUMBER_curtail':bussestrn3[1], 'TONUMBER2_curtail':bussestrn3[2]}
		PostCurtailmentProject_df = pd.DataFrame(data)


	if len(PostProject_df)>=1 or len(PreProject_df)>=1:
		merged_df = pd.merge(PostProject_df, PreProject_df, on=['ID','FROMNUMBER','TONUMBER','TONUMBER2'], how='left', suffixes=('_x', '_y'))
		merged_df = merged_df.rename(columns={'LOADING_x': 'LOADING_PostProject', 'LOADING_y': 'LOADING_PreProject', 'POWERFLOW_x': 'POWERFLOW_PostProject', 'POWERFLOW_y': 'POWERFLOW_PreProject', 'MW_x': 'MW_PostProject','MW_y': 'MW_PreProject' })
		merged_df = pd.merge(merged_df, PostCurtailmentProject_df, on=['ID','FROMNUMBER','TONUMBER','TONUMBER2'], how='left', suffixes=('_x', '_y'))
		# merged_df = pd.merge(merged_df, PostCurtailmentProject_df, on=['ID'], how='left', suffixes=('_x', '_y'))
		merged_df = merged_df.rename(columns={'LOADING': 'LOADING_PostCurtailment','POWERFLOW':'POWERFLOW_PostCurtailment'})    
		result2 = merged_df.loc[(merged_df['LOADING_PostProject']>Limit_Check) | (merged_df['LOADING_PreProject']>Limit_Check)| (merged_df['LOADING_PostCurtailment']>Limit_Check)]
		results = pd.concat([results, result2], ignore_index=True)
		
		# round all the values to 2 decimals 
		for col in results.columns:
			if pd.api.types.is_numeric_dtype(results[col]):
				results[col] = results[col].round(2)
		results['StudyCase'] = case
		return results , merged_volt_df
	else:
		results['StudyCase'] = case
		return results,merged_volt_df

















def BranchOverloads_v2(dirname, case, pre_case, arealist, cons, logger , SaveCases, Limit_Check): # this is a modified version of BranchOverloads which can populate df for Category A violations in which pre and post are compared against each other
	
	if logger is not None: logger.info("Saving the case with no contingenciy")
	psspy.case(os.path.join(dirname, case))
	Solve()
	if not Solve():
		if not logger is None: logger.error("The case is not solved!")
	ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
	if 'WP' in case:
		flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEB", "RATEB", "RATE4","P"])[1]
		flows1  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEB", "RATEB", "RATE3","P"])[1]
		flows[2] = [rate4 if rate4 > 0 else rate3 for rate4, rate3 in zip(flows[2], flows1[2])]  
	else:
		flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEA", "RATEA", "RATE3","P"])[1]
	busses = psspy.abrnint(0, 1, 3, 3, 1, ["FROMNUMBER", "TONUMBER"])[1]
	ids = psspy.abrnchar(0, 1, 3, 3, 1, ["ID", "FROMNAME", "TONAME"])[1]
	names = [psspy.brnnam(busses[0][i], busses[1][i], f"{ids[0][i]}")[1].replace(" ", "") for i in range(len(ids[0]))]
	actualFlows = [round(flows[0][i]*flows[1][i]/100.0,2) for i in range(len(ids[0]))]
	if cons==None:
		data = {'FROMNAME': ids[1][:],'TONAME': ids[2][:], 'ID': ids[0][:], 'NAME':names, 'NORMALRATING':flows[1],'POWERFLOW':actualFlows,'LOADING': flows[0], 'FROMNUMBER':busses[0], 'TONUMBER':busses[1], 'MW': flows[3]}
	else:
		data = {'FROMNAME': ids[1][:],'TONAME': ids[2][:], 'ID': ids[0][:], 'NAME':names, 'NORMALRATING':flows[1],'EmergencyRating':flows[2],'POWERFLOW':actualFlows,'LOADING': flows[0], 'FROMNUMBER':busses[0], 'TONUMBER':busses[1], 'MW': flows[3]}
	PostProject_df = pd.DataFrame(data)
	PostProject_Volt_df = VoltageValues(arealist, False, 'No Outage')
	

	PreProject_df = pd.DataFrame(columns=['ID','FROMNUMBER','TONUMBER','LOADING','POWERFLOW', 'MW'])
	PreProject_Volt_df = pd.DataFrame(columns=[	'Bus Number',	'Base Voltage (kV)',	'Voltage (kV)-pre',	'Voltage (PU)-pre'])
	volt_df = pd.DataFrame()
	if not pre_case is None:    
		psspy.case(os.path.join(dirname, pre_case))
		Solve()
		if not Solve():
			if not logger is None: logger.error(f"The case is not solved! for {pre_case}")        
		ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
		if 'WP' in case:
			flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEB", "RATEB", "RATE4","P"])[1]
			flows1  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEB", "RATEB", "RATE3","P"])[1]
			flows[2] = [rate4 if rate4 > 0 else rate3 for rate4, rate3 in zip(flows[2], flows1[2])]             
		else:
			flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEA", "RATEA", "RATE3","P"])[1]


		busses = psspy.abrnint(0, 1, 3, 3, 1, ["FROMNUMBER", "TONUMBER"])[1]
		ids = psspy.abrnchar(0, 1, 3, 3, 1, ["ID", "FROMNAME", "TONAME"])[1]
		names = [psspy.brnnam(busses[0][i], busses[1][i], f"{ids[0][i]}")[1].replace(" ", "") for i in range(len(ids[0]))]
		actualFlows = [round(flows[0][i]*flows[1][i]/100.0,2) for i in range(len(ids[0]))]
		data = {'ID': ids[0][:],'POWERFLOW':actualFlows,'LOADING': flows[0], 'FROMNUMBER':busses[0], 'TONUMBER':busses[1], 'MW':flows[3], 'FROMNUMBER_Pre':busses[0], 'TONUMBER_Pre':busses[1]}
		PreProject_df = pd.DataFrame(data)
		PreProject_Volt_df = VoltageValues(arealist, True, 'No Outage')
	volt_df = VoltageCehck(PreProject_Volt_df, PostProject_Volt_df, cons)



	merged_df = pd.merge(PostProject_df, PreProject_df, on=['ID','FROMNUMBER','TONUMBER'], how='left', suffixes=('_x', '_y'))
	merged_df = merged_df.rename(columns={'LOADING_x': 'LOADING_PostProject', 'LOADING_y': 'LOADING_PreProject', 'POWERFLOW_x': 'POWERFLOW_PostProject', 'POWERFLOW_y': 'POWERFLOW_PreProject','MW_x': 'MW_PostProject','MW_y': 'MW_PreProject' })
	results =  merged_df.loc[(merged_df['LOADING_PostProject'].fillna(0) > Limit_Check) | (merged_df['LOADING_PreProject'].fillna(0) > Limit_Check)]
	
		


	psspy.case(os.path.join(dirname, case))
	Solve()
	if not Solve():
		if not logger is None: logger.error("The case is not solved!")        

	if 'WP' in case:
		flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEB", "RATEB", "RATE4","P"])[1]
		flowstrn3_1  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEB", "RATEB", "RATE3","P"])[1]
		flowstrn3[2] = [rate4 if rate4 > 0 else rate3 for rate4, rate3 in zip(flowstrn3[2], flowstrn3_1[2])]                   
	else:
		flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEA", "RATEA", "RATE3","P"])[1]
	bussestrn3 = psspy.awndint(0, 1, 3, 1, 1,  ["WIND1NUMBER", "WIND2NUMBER","WIND3NUMBER", "WNDNUM"])[1]
	idstrn3 = psspy.awndchar(0, 1, 3, 1, 1,  ["ID","WIND1NAME","WIND2NAME","WIND3NAME"])[1]
	names = [psspy.tr3nam(bussestrn3[0][i], bussestrn3[1][i],bussestrn3[2][i], f"{idstrn3[0][i]}")[1].replace(" ", "") for i in range(len(idstrn3[0]))]
	actualFlows = [round(flowstrn3[0][i]*flowstrn3[1][i]/100.0,2) for i in range(len(idstrn3[0]))]
	if cons==None:
		data = {'FROMNAME': idstrn3[1][:],'TONAME': idstrn3[2][:], 'TONAME2': idstrn3[3][:], 'ID': idstrn3[0][:], 'NAME':names, 'NORMALRATING':flowstrn3[1],'POWERFLOW':actualFlows,'LOADING': flowstrn3[0], 'FROMNUMBER':bussestrn3[0], 'TONUMBER':bussestrn3[1], 'TONUMBER2':bussestrn3[2], 'MW':flowstrn3[3]}
	else:
		data = {'FROMNAME': idstrn3[1][:],'TONAME': idstrn3[2][:], 'TONAME2': idstrn3[3][:], 'ID': idstrn3[0][:], 'NAME':names, 'NORMALRATING':flowstrn3[1], 'EmergencyRating': flowstrn3[2], 'POWERFLOW':actualFlows,'LOADING': flowstrn3[0], 'FROMNUMBER':bussestrn3[0], 'TONUMBER':bussestrn3[1], 'TONUMBER2':bussestrn3[2], 'MW': flowstrn3[3]}
	PostProject_df = pd.DataFrame(data)

	PreProject_df = pd.DataFrame(columns=['ID','FROMNUMBER','TONUMBER','TONUMBER2','LOADING','POWERFLOW', "MW"])
	if not pre_case is None:    
		psspy.case(os.path.join(dirname, pre_case))
		Solve()
		if not Solve():
			if not logger is None: 
				logger.error(f"The case is not solved for {pre_case}!")            
		ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
		if 'WP' in case:
			flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEB", "RATEB","P"])[1]
		else:
			flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEA", "RATEA","P"])[1]
		bussestrn3 = psspy.awndint(0, 1, 3, 1, 1,  ["WIND1NUMBER", "WIND2NUMBER","WIND3NUMBER", "WNDNUM"])[1]
		idstrn3 = psspy.awndchar(0, 1, 3, 1, 1,  ["ID","WIND1NAME","WIND2NAME","WIND3NAME"])[1]
		names = [psspy.tr3nam(bussestrn3[0][i], bussestrn3[1][i],bussestrn3[2][i], f"{idstrn3[0][i]}")[1].replace(" ", "") for i in range(len(idstrn3[0]))]
		actualFlows = [round(flowstrn3[0][i]*flowstrn3[1][i]/100.0,2) for i in range(len(idstrn3[0]))]
		data = {'ID': idstrn3[0], 'POWERFLOW':actualFlows,'LOADING': flowstrn3[0], 'FROMNUMBER':bussestrn3[0], 'TONUMBER':bussestrn3[1], 'TONUMBER2':bussestrn3[2],"MW":flowstrn3[2]}
		PreProject_df = pd.DataFrame(data)
	

	if (PostProject_df[(PostProject_df != 'N/A').all(axis=1)].shape[0]>=1) and  (PreProject_df[(PreProject_df != 'N/A').all(axis=1)].shape[0]>=1):
		PostProject_df['index'] = PostProject_df.groupby(['ID','FROMNUMBER','TONUMBER','TONUMBER2']).cumcount() + 1
		PreProject_df['index'] = PreProject_df.groupby(['ID','FROMNUMBER','TONUMBER','TONUMBER2']).cumcount() + 1				
		merged_df = pd.merge(PostProject_df, PreProject_df, on=['ID','FROMNUMBER','TONUMBER','TONUMBER2', 'index'], how='left', suffixes=('_x', '_y'),)
		merged_df.drop(columns=['index'], inplace=True)
		merged_df = merged_df.rename(columns={'LOADING_x': 'LOADING_PostProject', 'LOADING_y': 'LOADING_PreProject', 'POWERFLOW_x': 'POWERFLOW_PostProject', 'POWERFLOW_y': 'POWERFLOW_PreProject','MW_x': 'MW_PostProject','MW_y': 'MW_PreProject' })				
	elif (PostProject_df[(PostProject_df != 'N/A').all(axis=1)].shape[0]>=1) and  (PreProject_df[(PreProject_df != 'N/A').all(axis=1)].shape[0]<1):
		merged_df = PostProject_df.rename(columns={'LOADING':'LOADING_PostProject', 'POWERFLOW':'POWERFLOW_PostProject','MW':'MW_PostProject'})
		merged_df = pd.merge(merged_df, PreProject_df, on=['ID','FROMNUMBER','TONUMBER','TONUMBER2'], how='left', suffixes=('_x', '_y'),)
		merged_df = pd.concat([merged_df, pd.DataFrame(columns=['LOADING_PreProject', 'POWERFLOW_PreProject', 'MW_PreProject'])], axis=1)
	elif (PostProject_df[(PostProject_df != 'N/A').all(axis=1)].shape[0]<1) and  (PreProject_df[(PreProject_df != 'N/A').all(axis=1)].shape[0]>=1):
		merged_df = PostProject_df.rename(columns={'LOADING':'LOADING_PreProject', 'POWERFLOW':'POWERFLOW_PreProject','MW':'MW_PreProject'})
		merged_df = pd.concat([merged_df, pd.DataFrame(columns=['LOADING_PostProject', 'POWERFLOW_PostProject', 'MW_PostProject'])], axis=1)				
		
	
	result2 =  merged_df.loc[(merged_df['LOADING_PostProject'].fillna(0) > Limit_Check) | (merged_df['LOADING_PreProject'].fillna(0) > Limit_Check)]
	results = pd.concat([results, result2], ignore_index=True)

	merged_volt_df = volt_df
	# round all the values to 2 decimals 
	for col in results.columns:
		if pd.api.types.is_numeric_dtype(results[col]):
			results[col] = results[col].round(2)
	
	if cons!=None:
		if not os.path.exists(os.path.join(dirname, "Cases")):
			os.makedirs(os.path.join(dirname, "Cases"))

		if logger is not None: logger.info(f"""The contingency cases will be saved in {os.path.join(dirname, "Cases")}""")
		results["Contingency"] = "No Outage"
		for con in cons:
			if logger is not None:  logger.info(f"Working on {con}.")
			psspy.case(os.path.join(dirname, case))
			psspy.solution_parameters_5([_i,100,_i,_i,10,_i,_i,20,0],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
			applyCon(cons[con], logger)
			extension = con.replace("(", "").replace(")", "").replace("'", '').replace('"', '').replace(" ","") #extension of the name for the saved case
			if logger is not None: logger.info(f"""Saving the case for {extension} Contingency as {case.replace(".sav", "_{0}.sav".format(extension))}""")
			
			psspy.save(os.path.join(dirname, "Cases", case.replace(".sav", "_{0}.sav".format(extension))))
			if SaveCases==1:  
				psspy.save(os.path.join(dirname, "Cases", case.replace(".sav", "_V35_{0}.sav".format(extension))))
			
			if Solve():
				PreProject_Volt_df = pd.DataFrame(columns=[	'Bus Number',	'Base Voltage (kV)',	'Voltage (kV)-pre',	'Voltage (PU)-pre'])
				PostProject_Volt_df = VoltageValues(arealist, False, con)			
				psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
				if 'WP' in case:
					flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEB", "RATEB", "RATE4","P"])[1]
					flows1  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEB", "RATEB", "RATE3","P"])[1]
					flows[2] = [rate4 if rate4 > 0 else rate3 for rate4, rate3 in zip(flows[2], flows1[2])]                     
				else:
					flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEA", "RATEA", "RATE3","P"])[1]
				busses = psspy.abrnint(0, 1, 3, 3, 1, ["FROMNUMBER", "TONUMBER"])[1]
				ids = psspy.abrnchar(0, 1, 3, 3, 1, ["ID", "FROMNAME", "TONAME"])[1]
				names = [psspy.brnnam(busses[0][i], busses[1][i], f"{ids[0][i]}")[1].replace(" ", "") for i in range(len(ids[0]))]
				actualFlows = [round(flows[0][i]*flows[1][i]/100.0,2) for i in range(len(ids[0]))]
				data = {'FROMNAME': ids[1][:],'TONAME': ids[2][:], 'ID': ids[0][:], 'NAME':names, 'NORMALRATING':flows[1],'EmergencyRating':flows[2],'POWERFLOW':actualFlows,'LOADING': flows[0], 'FROMNUMBER':busses[0], 'TONUMBER':busses[1], "Contingency": con, "StudyCase":case,"MW":flows[3]}
				PostProject_df = pd.DataFrame(data)
				PreProject_Volt_df = VoltageValues(arealist, True, con)
			else:             
				if not logger is None:
					logger.error(f'''The case is not solved for {case.replace(".sav", "_{0}".format(extension))}! The case is saved for further investigations.''')
					psspy.case(os.path.join(dirname, case))
					applyCon(cons[con], logger)
					if SaveCases==1:  
						psspy.save(os.path.join(dirname, "Cases", case.replace(".sav", "_V35_{0}Failed.sav".format(extension))))					                
					data = {'FROMNAME': "N/A",'TONAME': "N/A", 'ID': "N/A", 'NAME':"N/A", 'NORMALRATING':"N/A",'EmergencyRating':"N/A",'POWERFLOW':"N/A",'LOADING':10000, 'FROMNUMBER':"N/A", 'TONUMBER':"N/A", "Contingency": con, "StudyCase":case,"MW":"N/A"}
					PostProject_df = pd.DataFrame(data, index=[0])
			
			if not pre_case is None:    
				psspy.case(os.path.join(dirname, pre_case))
				applyCon(cons[con], logger)
				psspy.save(os.path.join(dirname, "Cases", pre_case.replace(".sav", "_{0}.sav".format(extension))))
				if SaveCases==1:  # only save in v34 if the user checked the SaveCases box
					psspy.save(os.path.join(dirname, "Cases", pre_case.replace(".sav", "_V35_{0}.sav".format(extension))))					
				if Solve():
					PreProject_Volt_df = VoltageValues(arealist, True, con)
					volt_df = VoltageCehck(PreProject_Volt_df, PostProject_Volt_df, cons)	
					merged_volt_df = pd.concat([merged_volt_df, volt_df], ignore_index=True)					
					ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
					if 'WP' in case:
						flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEB", "RATEB", "RATE4","P"])[1]
						flows1  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEB", "RATEB", "RATE3","P"])[1]
						flows[2] = [rate4 if rate4 > 0 else rate3 for rate4, rate3 in zip(flows[2], flows1[2])]                         
					else:
						flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEA", "RATEA", "RATE3","P"])[1]
					busses = psspy.abrnint(0, 1, 3, 3, 1, ["FROMNUMBER", "TONUMBER"])[1]
					ids = psspy.abrnchar(0, 1, 3, 3, 1, ["ID", "FROMNAME", "TONAME"])[1]
					names = [psspy.brnnam(busses[0][i], busses[1][i], f"{ids[0][i]}")[1].replace(" ", "") for i in range(len(ids[0]))]
					actualFlows = [round(flows[0][i]*flows[1][i]/100.0,2) for i in range(len(ids[0]))]
					data = {'ID': ids[0][:],'POWERFLOW':actualFlows,'LOADING': flows[0], 'FROMNUMBER':busses[0], 'TONUMBER':busses[1], 'MW':flows[3]}
					PreProject_df = pd.DataFrame(data)
				else:
					if not logger is None: 
						logger.error(f'''The case is not solved for {pre_case.replace(".sav", "_{0}".format(extension))}! The case is saved for further investigations.''')

						# print("h")  

					psspy.case(os.path.join(dirname, pre_case))
					applyCon(cons[con], logger)
					psspy.save(os.path.join(dirname, "Cases", pre_case.replace(".sav", "_V35_{0}Failed.sav".format(extension))))
					saveCase(os.path.join(dirname, "Cases"), pre_case.replace(".sav", "_V35_{0}Failed.sav".format(extension)), logger, False) 
										
					data = {'ID': "N/A",'POWERFLOW':"N/A",'LOADING': 10000, 'FROMNUMBER':"N/A", 'TONUMBER':"N/A",'MW':"N/A"}
					PreProject_df = pd.DataFrame(data, index=[0])
			
			if not pre_case is None:
				if len(PostProject_df)>1 and PreProject_df[(PreProject_df != 'N/A').all(axis=1)].shape[0]>=1:
					merged_df = pd.merge(PostProject_df, PreProject_df, on=['ID','FROMNUMBER','TONUMBER'], how='left', suffixes=('_x', '_y'))
				else:
					data = {'ID': "N/A",'POWERFLOW':"N/A",'LOADING': 10000, 'FROMNUMBER':"N/A", 'TONUMBER':"N/A",'MW':"N/A"}
					PreProject_df = pd.DataFrame(data, index=[0])  
					data = {'FROMNAME': "N/A",'TONAME': "N/A", 'ID': "N/A", 'NAME':"N/A", 'NORMALRATING':"N/A",'EmergencyRating':"N/A",'POWERFLOW':"N/A",'LOADING':10000, 'FROMNUMBER':"N/A", 'TONUMBER':"N/A", "Contingency": con, "StudyCase":case, "MW":"N/A"}
					PostProject_df = pd.DataFrame(data, index=[0])                
					merged_df = pd.merge(PostProject_df, PreProject_df, on=['ID','FROMNUMBER','TONUMBER'], how='left', suffixes=('_x', '_y')) 
				merged_df = merged_df.rename(columns={'LOADING_x': 'LOADING_PostProject', 'LOADING_y': 'LOADING_PreProject', 'POWERFLOW_x': 'POWERFLOW_PostProject', 'POWERFLOW_y': 'POWERFLOW_PreProject','MW_x': 'MW_PostProject','MW_y': 'MW_PreProject' })
			else:
				merged_df = PostProject_df
				merged_df = merged_df.rename(columns={'LOADING': 'LOADING_PostProject', 'POWERFLOW': 'POWERFLOW_PostProject','MW': 'MW_PostProject' })
				merged_df['LOADING_PreProject'] = 0.0
			try:
				result2 =  merged_df.loc[(merged_df['LOADING_PostProject'].fillna(0) > Limit_Check) | (merged_df['LOADING_PreProject'].fillna(0) > Limit_Check)]
			except:
				result2 =  merged_df.loc[merged_df['LOADING_PostProject']=='N/A']
			results = pd.concat([results, result2], ignore_index=True)
				
					
			
			psspy.case(os.path.join(dirname, "Cases", case.replace(".sav", "_{0}.sav".format(extension))))
			Solve()
			if not Solve():
				if not logger is None: logger.error(f'''The {case.replace(".sav", "_{0}.sav".format(extension))} is not solved!''')   
				data = {'FROMNAME': "N/A",'TONAME': "N/A", 'TONAME2':"N/A", 'ID': "N/A", 'NAME':"N/A", 'NORMALRATING':"N/A",'EmergencyRating':"N/A", 'POWERFLOW':"N/A",'LOADING': 10000, 'FROMNUMBER':"N/A", 'TONUMBER':"N/A", 'TONUMBER2':"N/A", "Contingency": con, "MW":"N/A"}
				PostProject_df = pd.DataFrame(data, index=[0])     
			else:                   

				if 'WP' in case:
					flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEB", "RATEB", "RATE4","P"])[1]
				else:
					flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEA", "RATEA", "RATE3","P"])[1]
				bussestrn3 = psspy.awndint(0, 1, 3, 1, 1,  ["WIND1NUMBER", "WIND2NUMBER","WIND3NUMBER", "WNDNUM"])[1]
				idstrn3 = psspy.awndchar(0, 1, 3, 1, 1,  ["ID","WIND1NAME","WIND2NAME","WIND3NAME"])[1]
				names = [psspy.tr3nam(bussestrn3[0][i], bussestrn3[1][i],bussestrn3[2][i], f"{idstrn3[0][i]}")[1].replace(" ", "") for i in range(len(idstrn3[0]))]
				try:
					actualFlows = [round(flowstrn3[0][i]*flowstrn3[1][i]/100.0,2) for i in range(len(idstrn3[0]))]
					data = {'FROMNAME': idstrn3[1][:],'TONAME': idstrn3[2][:], 'TONAME2': idstrn3[3][:], 'ID': idstrn3[0][:], 'NAME':names, 'NORMALRATING':flowstrn3[1],'EmergencyRating':flowstrn3[2], 'POWERFLOW':actualFlows,'LOADING': flowstrn3[0], 'FROMNUMBER':bussestrn3[0], 'TONUMBER':bussestrn3[1], 'TONUMBER2':bussestrn3[2], "Contingency": con, "MW":flowstrn3[2]}
					
					PostProject_df = pd.DataFrame(data)
				except:
					data = {'FROMNAME': "N/A",'TONAME': "N/A", 'TONAME2':"N/A", 'ID': "N/A", 'NAME':"N/A", 'NORMALRATING':"N/A",'EmergencyRating':"N/A", 'POWERFLOW':"N/A",'LOADING': 10000, 'FROMNUMBER':"N/A", 'TONUMBER':"N/A", 'TONUMBER2':"N/A", "Contingency": con, "MW":"N/A"}
					PostProject_df = pd.DataFrame(data, index=[0])    


			PreProject_df = pd.DataFrame(columns=['ID','FROMNUMBER','TONUMBER','TONUMBER2','LOADING','POWERFLOW','MW'])
			if not pre_case is None:    
				psspy.case(os.path.join(dirname, "Cases", pre_case.replace(".sav", "_{0}.sav".format(extension))))
				Solve()
				if not Solve():
					if not logger is None: 
						logger.error(f'''The case is not solved for {pre_case.replace(".sav", "_{0}.sav".format(extension))}!''')  
					data = {'ID': "N/A", 'POWERFLOW':"N/A",'LOADING': "N/A", 'FROMNUMBER':"N/A", 'TONUMBER':"N/A", 'TONUMBER2':"N/A", "MW":"N/A"}
					PreProject_df = pd.DataFrame(data, index=[0])
				else:          
					ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
					if 'WP' in case:
						flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEB", "RATEB","P"])[1]
					else:
						flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEA", "RATEA","P"])[1]
					bussestrn3 = psspy.awndint(0, 1, 3, 1, 1,  ["WIND1NUMBER", "WIND2NUMBER","WIND3NUMBER", "WNDNUM"])[1]
					idstrn3 = psspy.awndchar(0, 1, 3, 1, 1,  ["ID","WIND1NAME","WIND2NAME","WIND3NAME"])[1]
					names = [psspy.tr3nam(bussestrn3[0][i], bussestrn3[1][i],bussestrn3[2][i], f"{idstrn3[0][i]}")[1].replace(" ", "") for i in range(len(idstrn3[0]))]
					actualFlows = [round(flowstrn3[0][i]*flowstrn3[1][i]/100.0,2) for i in range(len(idstrn3[0]))]
					data = {'ID': idstrn3[0][:], 'POWERFLOW':actualFlows,'LOADING': flowstrn3[0], 'FROMNUMBER':bussestrn3[0], 'TONUMBER':bussestrn3[1], 'TONUMBER2':bussestrn3[2], "MW":flowstrn3[2]}
					PreProject_df = pd.DataFrame(data)
			if (PostProject_df[(PostProject_df != 'N/A').all(axis=1)].shape[0]>=1) and  (PreProject_df[(PreProject_df != 'N/A').all(axis=1)].shape[0]>=1):
				PostProject_df['index'] = PostProject_df.groupby(['ID','FROMNUMBER','TONUMBER','TONUMBER2']).cumcount() + 1
				PreProject_df['index'] = PreProject_df.groupby(['ID','FROMNUMBER','TONUMBER','TONUMBER2']).cumcount() + 1				
				merged_df = pd.merge(PostProject_df, PreProject_df, on=['ID','FROMNUMBER','TONUMBER','TONUMBER2', 'index'], how='left', suffixes=('_x', '_y'),)
				merged_df.drop(columns=['index'], inplace=True)
				merged_df = merged_df.rename(columns={'LOADING_x': 'LOADING_PostProject', 'LOADING_y': 'LOADING_PreProject', 'POWERFLOW_x': 'POWERFLOW_PostProject', 'POWERFLOW_y': 'POWERFLOW_PreProject','MW_x': 'MW_PostProject','MW_y': 'MW_PreProject' })				
			elif (PostProject_df[(PostProject_df != 'N/A').all(axis=1)].shape[0]>=1) and  (PreProject_df[(PreProject_df != 'N/A').all(axis=1)].shape[0]<1):
				merged_df = PostProject_df.rename(columns={'LOADING':'LOADING_PostProject', 'POWERFLOW':'POWERFLOW_PostProject','MW':'MW_PostProject'})
				merged_df = pd.concat([merged_df, pd.DataFrame(columns=['LOADING_PreProject', 'POWERFLOW_PreProject', 'MW_PreProject'])], axis=1)
				merged_df['LOADING_PreProject'] = merged_df['LOADING_PreProject'].astype(float)
			elif (PostProject_df[(PostProject_df != 'N/A').all(axis=1)].shape[0]<1) and  (PreProject_df[(PreProject_df != 'N/A').all(axis=1)].shape[0]>=1):
				merged_df = PostProject_df.rename(columns={'LOADING':'LOADING_PreProject', 'POWERFLOW':'POWERFLOW_PreProject','MW':'MW_PreProject'})
				merged_df = pd.concat([merged_df, pd.DataFrame(columns=['LOADING_PostProject', 'POWERFLOW_PostProject', 'MW_PostProject'])], axis=1)				

			try:
				result2 =  merged_df.loc[(merged_df['LOADING_PostProject'].fillna(0) > Limit_Check) | (merged_df['LOADING_PreProject'].fillna(0) > Limit_Check)]
			except:
				result2 =  merged_df.loc[merged_df['LOADING_PostProject']=='N/A']        
			results = pd.concat([results, result2], ignore_index=True)
			
			# round all the values to 2 decimals 
			for col in results.columns:
				if pd.api.types.is_numeric_dtype(results[col]):
					results[col] = results[col].round(2)                
	if os.path.exists(os.path.join(dirname, "Cases")):
		if os.path.isdir(os.path.join(dirname, "Cases")):
			for file in os.listdir(os.path.join(dirname, "Cases")):
				if ("V35_" not in file):
					os.remove(os.path.join(os.path.join(dirname, "Cases"), file))

	
	
	psspy.case(os.path.join(dirname, case))
	PostProject_Volt_df = VoltageValues(arealist, False, 'No Outage')
	PostProject_Volt_df['Voltage (kV)-post_initial'] = PostProject_Volt_df['Voltage (kV)-post']
	merged_volt_df = pd.merge(merged_volt_df,PostProject_Volt_df[['Voltage (kV)-post_initial','Bus Number']],how='left',on=['Bus Number'] )
	if pre_case: 
		psspy.case(os.path.join(dirname, pre_case))
		PreProject_Volt_df = VoltageValues(arealist, True, 'No Outage')	
		PreProject_Volt_df['Voltage (kV)-pre_initial'] = PreProject_Volt_df['Voltage (kV)-pre']
		merged_volt_df = pd.merge(merged_volt_df,PreProject_Volt_df[['Voltage (kV)-pre_initial','Bus Number']],how='left',on=['Bus Number'] )
	merged_volt_df['StudyCase'] = case
	results['StudyCase'] = case
	logger.info(f"All contingencies for {case} is done.")
	return results, merged_volt_df   




def BranchOverloads(dirname, case, arealist, Limit_Check, cons = None, logger = None):
	print(os.path.join(dirname, "Cases", case.replace(".sav", "_No_outage.sav")))
	if logger is not None: logger.info("Saving the case with no contingenciy")
	psspy.save(os.path.join(dirname, "Cases", case.replace(".sav", "_No_outage.sav")))
	ierr = psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
	if 'WP' in case:
		flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEB", "RATEB"])[1]
	else:
		flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEA", "RATEA"])[1]
	busses = psspy.abrnint(0, 1, 3, 3, 1, ["FROMNUMBER", "TONUMBER"])[1]
	ids = psspy.abrnchar(0, 1, 3, 3, 1, ["ID"])[1][0]
	results = []
	count = len(ids)
	
	for i in range(count):
		if flows[0][i] > Limit_Check:
			bus1 = busses[0][i]
			bus2 = busses[1][i]
			id = ids[i]
			name = psspy.brnnam(bus1, bus2, id)[1].strip()
			results.append([case, "No_outage", "N/A", name, str(bus1) + " " + str(bus2) + " " + id, round(flows[1][i], 2), round(flows[0][i], 2)])
	
	if 'WP' in case:
		flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEB", "RATEB"])[1]
	else:
		flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEA", "RATEA"])[1]
	bussestrn3 = psspy.awndint(0, 1, 3, 1, 1,  ["WIND1NUMBER", "WIND2NUMBER","WIND3NUMBER", "WNDNUM"])[1]
	idstrn3 = psspy.awndchar(0, 1, 3, 1, 1,  ["ID"])[1][0]
	counttrn3 = len(idstrn3)
	
	for i in range(counttrn3):
		if flowstrn3[0][i] > Limit_Check:
			bus1 = bussestrn3[0][i]
			bus2 = bussestrn3[1][i]
			bus3 = bussestrn3[2][i]
			id = idstrn3[i]
			wnd = bussestrn3[3][i]
			name = psspy.tr3nam(bus1, bus2, bus3, id)[1].strip()
			results.append([case, "No_outage", "N/A", "{0} winding {1}".format(name, wnd), str(bus1) + " " + str(bus2) + " " + str(bus3) + " " + id, round(flowstrn3[1][i], 2), round(flowstrn3[0][i], 2)])
	
	progresscount = 0
	try:
		progresstotal = len(cons)
	except TypeError:
		progresstotal = 1
	
	if cons != None:
		for con in cons:
			print("Running power flow for {0} contingency".format([con]))
			if logger is not None: logger.info("Running power flow for {0} contingency".format([con]))
			psspy.case(os.path.join(dirname,case))
			psspy.solution_parameters_5([_i,100,_i,_i,10,_i,_i,20,0],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
			for conEntry in cons[con]:
				if conEntry[0] == "disconnect":
					psspy.dscn(conEntry[1])
				if conEntry[0] == "brn":
					psspy.progress_output(6,"",[])
					psspy.branch_chng_3(conEntry[1],conEntry[2],conEntry[3],[0,_i,_i,_i,_i,_i],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],[ _f, _f, _f, _f,_f,_f,_f,_f,_f,_f,_f,_f],_s)
					psspy.two_winding_chng_6(conEntry[1],conEntry[2],conEntry[3],[0,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i,_i],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],[ _f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],_s,_s)
					psspy.progress_output(2,os.path.join(dirname, "psse_output.log"),[])
				if conEntry[0] == "3wnd":
					psspy.three_wnd_imped_chng_4(conEntry[1],conEntry[2],conEntry[3],conEntry[4],[_i,_i,_i,_i,_i,_i,_i,0,_i,_i,_i,_i,_i],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f],_s,_s)
				if conEntry[0] == "mach":
					print( conEntry)
					psspy.machine_chng_2(conEntry[1],conEntry[2],[0,_i,_i,_i,_i,_i],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])
			Solve()
			extension = con.replace("(", "").replace(")", "").replace("'", "").replace(" ","_") #extension of the name for the saved case
			psspy.save(os.path.join(dirname, "Cases", case.replace(".sav", "_{0}.sav".format(extension))))
			if Solve():
				psspy.bsys(0,1,[ 60, 500.],len(arealist),arealist,0,[],0,[],0,[])
				if 'WP' in case:
					flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEB", "RATEB"])[1]
				else:
					flows  = psspy.abrnreal(0, 1, 3, 3, 1, ["MAXPCTRATEA", "RATEA"])[1]
				busses = psspy.abrnint(0, 1, 3, 3, 1, ["FROMNUMBER", "TONUMBER"])[1]
				ids = psspy.abrnchar(0, 1, 3, 3, 1, ["ID"])[1][0]
				count = len(ids)
				
				for i in range(count):
					if flows[0][i] > Limit_Check:
						bus1 = busses[0][i]
						bus2 = busses[1][i]
						id = ids[i]
						name = psspy.brnnam(bus1, bus2, id)[1].strip()
						results.append([case, con, cons[con], name, str(bus1) + " " + str(bus2) + " " + str(id), round(flows[1][i], 2), round(flows[0][i], 2)])
				
				if 'WP' in case:
					flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEB", "RATEB"])[1]
				else:
					flowstrn3  = psspy.awndreal(0, 1, 3, 1, 1,  ["MAXPCTRATEA", "RATEA"])[1]
				bussestrn3 = psspy.awndint(0, 1, 3, 1, 1,  ["WIND1NUMBER", "WIND2NUMBER","WIND3NUMBER", "WNDNUM"])[1]
				idstrn3 = psspy.awndchar(0, 1, 3, 1, 1,  ["ID"])[1][0]
				counttrn3 = len(idstrn3)
				
				for i in range(counttrn3):
					if flowstrn3[0][i] > Limit_Check:
						bus1 = bussestrn3[0][i]
						bus2 = bussestrn3[1][i]
						bus3 = bussestrn3[2][i]
						id = idstrn3[i]
						wnd = bussestrn3[3][i]
						name = psspy.tr3nam(bus1, bus2, bus3, id)[1].strip()
						results.append([case, con, cons[con], "{0} winding {1}".format(name, wnd), str(bus1) + " " + str(bus2) + " " + str(bus3) + " " + id, round(flowstrn3[1][i], 2), round(flowstrn3[0][i], 2)])
			else:
				if logger is not None: logger.error("Case failed to solve for {0} contingency".format([con]))
				results.append([case, con, cons[con], "ALL", "ALL", "ALL", "Case failed to solve"])
	
	   
	return results
	
	
def CreateCombinations(inputxlsx):
	projectalts = defaultdict(list)
	combinations = []
	commonbusses = dict()
	altbusses = dict()
	
	print( inputxlsx)
	
	for row in inputxlsx[1:]:
		if row[2] == "Y":
			commonbusses[row[0]] = row[9].split(" ,")
		else:
			projectalts[row[0]].append(int(row[3]))
			altbusses["{0}-{1}".format(row[0], row[3])] = row[9].split(" ,")
			
	for key in projectalts:
		print( key, projectalts[key])
	for key in commonbusses:
		print( key, commonbusses[key])
	for key in altbusses:
		print( key, altbusses[key])
	
	return combinations, commonbusses, altbusses
	
## Main Function ##
if __name__ == "__main__":
	
	psspy.psseinit(16000) 
	_i=psspy.getdefaultint()
	_f=psspy.getdefaultreal()
	_s=psspy.getdefaultchar()
	#temp_file_path = sys.argv[1]

	temp_file_path = sys.argv[1]

# Read the arguments from the temporary file
	with open(temp_file_path, 'r') as temp_file:
		lines = temp_file.read().splitlines()

  
	try:
		run_local = 1  # change it to 1 to run it without requring input arguments from user
		if run_local==0:
			logger = MyLogger(r"T:\Rshared\Delivery Shared\Python Files\Sina\GitRepo\cas-automation-tools\CPAT GUI\Functions\Code_test\powerflow\New folder\New folder", "PowerFlow_{0}.log".format(datetime.now().strftime("%Y%m%d%H%M%S")))
			region = 'NortWest'
			logger.info("study areas are: {0}".format('SouthEast'))
			MasterSavedCase = ['S4_NW_2029_SL_HG_Post.sav', 'S4_NW_2029_SL_HG_Post-Post.sav']
			logger.info("Master base case is: {0}".format(MasterSavedCase))
			dirname = r'T:\Rshared\Delivery Shared\Python Files\Sina\GitRepo\cas-automation-tools\CPAT GUI\Functions\Code_test\powerflow\New folder\test_final'
			input_dir = r'T:\Rshared\Delivery Shared\Python Files\Sina\GitRepo\cas-automation-tools\CPAT GUI\Functions\Code_test\powerflow\New folder\test_final'
			# inputcsv = '20231020140007.csv'
			inputcsv = str('')
			confile = str('')
			regenrateShiftFactor = int('0')
			individulCase = int('1')
			runCon = eval('False')
			Limit_Check =99.99
			print('run_con', runCon)
			areas = eval('[43, 48, 47, 52, 54, 4]')
			logger.info("study areas are: {0} {1}".format(areas,type(areas)))
			print(os.path.join(inputcsv))
		else:
			# Access the parameters
			region = str(lines[0])
			logger = MyLogger(str(lines[2]), "PowerFlow_{0}.log".format(datetime.now().strftime("%Y%m%d%H%M%S")))
			logger.info("study areas are: {0}".format(region))
			MasterSavedCase = [s + '.sav' for s in lines[1].replace('[',"").replace(']',"").split(".sav")[:-1]]
			MasterSavedCase = [s.replace("'","") for s in MasterSavedCase]
			MasterSavedCase = [s.replace(",","") for s in MasterSavedCase]
			MasterSavedCase = [s.lstrip() for s in MasterSavedCase]
			MasterSavedCase = [s.rstrip() for s in MasterSavedCase]
			logger.info("Master base case is: {0}".format(MasterSavedCase))
			dirname = str(lines[2])
			input_dir = str(lines[3])
			inputcsv = str(lines[4])
			confile = str(lines[5])
			regenrateShiftFactor = not(eval(lines[6]))
			individulCase = int(lines[7])
			runCon = eval(lines[8])
			areas = eval(lines[9])
			GenerateSingleElementContingency = int(lines[10])
			SaveCases = int(lines[11])
			Limit_Check = float(lines[13])
			# logger = MyLogger(str(sys.argv[3]), "PowerFlow_{0}.log".format(datetime.now().strftime("%Y%m%d%H%M%S")))
			# region = str(sys.argv[1])
			# logger.info("study areas are: {0}".format(str(sys.argv[1])))
			# MasterSavedCase = [s + '.sav' for s in sys.argv[2].split(".sav")[:-1]]
			# logger.info("Master base case is: {0}".format(MasterSavedCase))
			# dirname = str(sys.argv[3])
			# input_dir = str(sys.argv[4])
			# inputcsv = str(sys.argv[5])
			# confile = str(sys.argv[6])
			# regenrateShiftFactor = int(sys.argv[7])
			# individulCase = int(sys.argv[8])
			# runCon = eval(sys.argv[9])
			# areas = eval(sys.argv[10])
			if not input_dir=="None":
				try:
					df = pd.read_csv(os.path.join(input_dir,sys.argv[11])) 
					logger.info(f"Customized input file is provided,")
				except:
					df = pd.DataFrame()
			else:
				df = pd.DataFrame()        
			logger.info("study areas are: {0} {1}".format(areas,type(areas)))
			print(os.path.join(inputcsv))
			try:
				GEFF_buses = []
				if os.path.exists(os.path.join(lines[3],lines[12])):
					cusom_Input_file_path = os.path.join(lines[3],lines[12])
					Input_df = pd.read_csv(cusom_Input_file_path)
					GEFF_buses = Input_df.loc[Input_df['sheet_name']=="GEFF"]
					GEFF_buses.dropna(axis = 0, how = 'all', inplace = True)
					GEFF_buses.dropna(axis = 1, how = 'all', inplace = True)		
			except:
					GEFF_buses = pd.DataFrame()
					pass	
		

		
		
		if runCon:
			masterCaseLookup = {} # this is to find the pre-project cases for each study case for reporting purposes.
			for case in MasterSavedCase:
				if len(set(MasterSavedCase).difference({case})):  # there are more than one cases try to find the pre_project case
					masterCaseLookup[case] = find_pre_project_case(set(MasterSavedCase).difference({case}),case)
				else:
					masterCaseLookup[case] = None
			# remove the cases which are the preproject case
			masterCaseLookup = {key: value for key, value in masterCaseLookup.items() if key not in masterCaseLookup.values()}
		else:
			print('sina3')
			masterCaseLookup = {} # this is to find the pre-project cases for each study case for reporting purposes.
			print('MasterSavedCase001:  ', set(MasterSavedCase))
			for case in MasterSavedCase:
				print('MasterSavedCase000:  ', set(MasterSavedCase).difference({case}))
				print('case: ', len(set(MasterSavedCase).difference({case})))
				if len(set(MasterSavedCase).difference({case})):  # there are more than one cases try to find the pre_project case

					masterCaseLookup[case] = find_related_cases(list(set(MasterSavedCase).difference({case})),case)
					print('masterCaseLookup11: ', masterCaseLookup)
				else:
					masterCaseLookup[case] = case_obj()
					print('masterCaseLookup22: ', masterCaseLookup)            
			# remove the cases which are the preproject case
			masterCaseLookup1 = {}
			for key1, value1 in masterCaseLookup.items(): 
				flag = 0
				if not(masterCaseLookup[key1].curtailedCase is None) and not(masterCaseLookup[key1].preCase is None): 

					for key2, value2 in masterCaseLookup.items():
						if key2!=key1 and (key1 == masterCaseLookup[key2].curtailedCase or key1 == masterCaseLookup[key2].preCase):
							flag = 1


				else:
					

					for key2, value2 in masterCaseLookup.items():
						if key2!=key1 and ( key1 == masterCaseLookup[key2].curtailedCase):
							flag = 1

				if flag==0:
					masterCaseLookup1[key1] = masterCaseLookup[key1]                    

			masterCaseLookup = masterCaseLookup1 
			print('masterCaseLookup:   ', masterCaseLookup)

		print("Dirname is ::::::::::::", dirname)

		if not inputcsv =='': 
			print('sina')   
			if not runCon:
				print(os.path.join(input_dir, inputcsv))
				inputDF = pd.read_csv(os.path.join(input_dir, inputcsv))
				inputDF = inputDF.loc[inputDF['Add?']=='Y']
				# if individulCase==1:
				#     combs = [comb for comb in combs if len(comb)==1]        
				combs = Combinations.CreateCombinations(inputDF, individulCase)
				print(combs)
				caseID = {}
				i = 0
				MasterSavedCase = MasterSavedCase[0]
				logger.info("Creating Alternatives")
				workingCase_sav_path =  os.path.join(dirname, MasterSavedCase)
				psspy.case(workingCase_sav_path)     
				psspy.solution_parameters_5([_i,100,_i,_i,10,_i,_i,20,0],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])     
				if not os.path.isdir(os.path.join(dirname, "BaseCases")):
					os.makedirs(os.path.join(dirname, "BaseCases"))
				for comb in combs:
					caseID[comb] = MasterSavedCase
					print(caseID[comb])
					caseID[comb] = caseID[comb].replace(".sav","_"+str(i))+'.sav'
					logger.info("Working on combination {0} which will be saved as {1}".format(comb, caseID[comb]))
					if comb==('P2778-Alt1',):
						print('yes')
					Combinations.createScenarios(comb, inputDF, dirname, MasterSavedCase, caseID[comb], logger, OpposingRegions[region])
					psspy.close_powerflow()
					psspy.psseinit(16000)   
					ierr = psspy.case(workingCase_sav_path)      
					psspy.solution_parameters_5([_i,100,_i,_i,10,_i,_i,20,0],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])        
					print("ierr", ierr)
					print("comb", caseID[comb])
					print(combs)
					i+=1
				df = pd.DataFrame(caseID.items(), columns=['Combination', 'Filename'])
				df.to_csv(os.path.join(dirname, 'reference.csv'))

				psspy.progress_output(2,os.path.join(dirname, "psse_output_{0}.log".format(datetime.now().strftime("%Y%m%d%H%M%S"))),[])
				print("...done!\n")

				# Thermalresults = []
				# Voltageresults = []
				# logger.info("Starting N-0 Power Flow Studies:")
				# if not os.path.isdir(os.path.join(dirname, "Cases")):
				#     os.makedirs(os.path.join(dirname, "Cases"))
				# for id in caseID:
				#     logger.info("Working on N-0 Power Flow for {0}".format(caseID[id]))
				#     full_file_path = os.path.join(dirname, "BaseCases",caseID[id])
				#     BaseCaseDir = os.path.join(dirname, "BaseCases")
				#     if os.path.isfile(full_file_path):
						
				#         saveCase(BaseCaseDir, caseID[id],logger)
				#         psspy.case(os.path.join(BaseCaseDir,caseID[id]))
				#         Thermalresults += (BranchOverloads(dirname,caseID[id], areas))
				#         Voltageresults += VoltageIssues(caseID[id], areas, BaseCaseDir)
				# logger.info("N-0 Power Flow is Completed")

				
			else:
				logger.info("Running Contingency Analysis")
				print("Dirname is ::::::::::::", dirname)
				Thermalresults = []
				# for file in os.listdir(dirname):
				file = MasterSavedCase
				
				err = psspy.case(os.path.join(dirname,file))
				print("***********",os.path.join(dirname,file), "###############", err)
				Thermalresults += (BranchOverloads(dirname,file, areas,Limit_Check))  
				cons, skips = buildConList(dirname, areas, confile)
				Thermalresults += (BranchOverloads(dirname, file, areas, cons, logger, Limit_Check))
		else:
			# print('sina2')
			if runCon:
				logger.info("Running Contingency Analysis")
				print("Dirname is ::::::::::::", dirname)
				Thermalresults  = pd.DataFrame()
				Voltageresults = pd.DataFrame()
				for case in masterCaseLookup: 
					cons, skips = buildConList(input_dir, confile, GenerateSingleElementContingency, areas, dirname, case, logger)
					Thermalresults1, Voltageresults1 = BranchOverloads_v2(dirname, case, masterCaseLookup[case], areas, cons, logger, SaveCases, Limit_Check) 
					Thermalresults = pd.concat([Thermalresults1, Thermalresults], ignore_index=True)
					Voltageresults = pd.concat([Voltageresults1, Voltageresults], ignore_index=True)
					logger.info("Contingency Analysis is Completed")                              
			else:
				cons = {}
				Thermalresults  = pd.DataFrame()
				ShiftFactor_df = pd.DataFrame()
				Voltageresults =pd.DataFrame()
				for case in masterCaseLookup:
					print('case in the loop:  ', masterCaseLookup[case])
					err = psspy.case(os.path.join(dirname,case))

					logger.info("Starting N-0 Power Flow Studies:")
					Thermalresults1, Voltageresults1 = BranchOverloads_v3(dirname,case, masterCaseLookup[case].preCase, masterCaseLookup[case].curtailedCase, areas,None, Limit_Check, logger)
					# Thermalresults1, Voltageresults1 = BranchOverloads_v3(dirname,case, None, None, areas,None, Limit_Check, logger)
					Thermalresults = pd.concat([Thermalresults1, Thermalresults], ignore_index=True)
					Voltageresults = pd.concat([Voltageresults1, Voltageresults], ignore_index=True)
					
				
					if len(Thermalresults1)>=1 and regenrateShiftFactor:
						logger.info(f"Creating shift factors for {case}")
						logger.info(Thermalresults1.columns)
						# psspy.psseinit(16000) 
						err = psspy.case(os.path.join(dirname,case))
						logger.info(os.path.join(dirname,case))
						try:
							SF = ShiftFactor()
							logger.info(Solve())
						except Exception as e:
							logger.info(f"Error occurred while running ShiftFactor(): {e}")
							# logging.error("Error occurred while running ShiftFactor()", exc_info=True)						

						SF.process('1', os.path.join(dirname,case),"ShiftFactor1")	
						ShiftFactor_df1 = pd.read_csv(os.path.join(dirname, "ShiftFactor1.csv"))  
						ShiftFactor_df1 = Thermalresults1.merge(ShiftFactor_df1,how='left', left_on = ['FROMNUMBER','TONUMBER','ID'], right_on = ['fromBus','toBus','branchId']) 
						ShiftFactor_df1 = ShiftFactor_df1[['FROMNAME','FROMNUMBER','TONUMBER','TONAME','BranchName', 'fromBus', 'toBus', 'ID', 'injectionBusName','injectionBus','genEffectivness', 'MW_PostProject', 'NORMALRATING','StudyCase']]
						ShiftFactor_df1['genEffectivness'] = np.sign(ShiftFactor_df1['MW_PostProject']) * ShiftFactor_df1['genEffectivness']
						ShiftFactor_df1['injectionBusName'] = ShiftFactor_df1.apply(replace_BusnNmeWithBusNumber, axis=1)
						# ShiftFactor_df1['StudyCase']=case
						ShiftFactor_df = pd.concat([ShiftFactor_df1, ShiftFactor_df], ignore_index=True)
						if os.path.exists(os.path.join(dirname, "ShiftFactor1.csv")):
							os.remove(os.path.join(dirname, "ShiftFactor1.csv"))   
						# ShiftFactor_df_pvt.to_csv(os.path.join(dirname, f'''ShiftFactor_CATA_{case.replace('.sav', '.csv')}'''), index=True)
				try:
					ShiftFactor_df_pvt = ShiftFactor_df.pivot_table(index=['FROMNAME','FROMNUMBER','TONAME','TONUMBER','ID', 'BranchName','StudyCase'], columns=['injectionBusName','injectionBus'], values='genEffectivness', aggfunc='mean')
				except:
					pass
				logger.info("N-0 Power Flow is Completed")
			try:
				if  runCon:
					if len(GEFF_buses)>0:
						Sensitivity_df(Thermalresults, dirname, logger, areas, cons, GEFF_buses)
					else:
						Sensitivity(Thermalresults, dirname, logger, areas, cons)
		

			except:
				logger.warning("Issue with caculating the shift factors for selective untis. (this is important only for CAT B)")




		# switch the from bus and to bus if the MW flow is opposite
		for index, row in Thermalresults.iterrows():
			try: 
				if Thermalresults.at[index, 'MW_PostProject']<0:
					fromname =  Thermalresults.at[index, 'FROMNAME']
					fromnumber = Thermalresults.at[index, 'FROMNUMBER']
					toname = Thermalresults.at[index, 'TONAME']
					tonumber = Thermalresults.at[index, 'TONUMBER']
					Thermalresults.at[index, 'FROMNAME'] = toname
					Thermalresults.at[index, 'FROMNUMBER'] = tonumber
					Thermalresults.at[index, 'TONAME'] = fromname
					Thermalresults.at[index, 'TONUMBER'] = fromnumber

				if Thermalresults.at[index, 'MW_PostCurtailment']<0:
					FROMNUMBER_curt = Thermalresults.at[index, 'FROMNUMBER_curtail']
					TONUMBER_curt = Thermalresults.at[index, 'TONUMBER_curtail']
					Thermalresults.at[index, 'FROMNUMBER_curtail'] = TONUMBER_curt
					Thermalresults.at[index, 'TONUMBER_curtail'] = FROMNUMBER_curt

				if Thermalresults.at[index, 'MW_PreProject']<0:
					FROMNUMBER_Pre = Thermalresults.at[index, 'FROMNUMBER_Pre']
					TONUMBER_Pre = Thermalresults.at[index, 'TONUMBER_Pre']
					Thermalresults.at[index, 'FROMNUMBER_Pre'] = TONUMBER_Pre
					Thermalresults.at[index, 'TONUMBER_Pre'] = FROMNUMBER_Pre
			except:
				pass
		


		filenum = 1
		if isinstance(Thermalresults, pd.DataFrame): 
			if runCon:
				while os.path.isfile(os.path.join(dirname, f"CatBViolationReport{filenum}.xlsm")):
					filenum = filenum+1 
				outputFile = f"CatBViolationReport{filenum}.xlsm"
				master_excel_file_path = r'T:\Rshared\Delivery Shared\Python Files\Sina\GitRepo\CatBViolationReportTemplate_Dev.xlsm'
			else:
				while os.path.isfile(os.path.join(dirname, f"CatAViolationReport{filenum}.xlsx")):
					filenum = filenum+1 
				outputFile = f"CatAViolationReport{filenum}.xlsx"
				master_excel_file_path = r'T:\Rshared\Delivery Shared\Python Files\Sina\GitRepo\CatAViolationReportTemplate_Dev.xlsx'
				# master_excel_file_path = r'T:\Rshared\Delivery Shared\Python Files\Sina\GitRepo\CatAViolationReportTemplate_Dev.xlsm'

			if len(Voltageresults)>0 or len(Thermalresults)>0:

				# Specify the existing Excel file path
				logger.info(f"Saving the result in {outputFile}.")
				excel_file_path = os.path.join(dirname,outputFile)
				shutil.copyfile(master_excel_file_path, excel_file_path)
				# Load the existing workbook
				if ".xlsm" in excel_file_path:
					wb = load_workbook(excel_file_path, keep_vba=True)
				else: wb = load_workbook(excel_file_path)

				# Access the sheet and clear its content
				ws = wb['Voltage_RAW']
				ws.delete_cols(1, 2000)  # Delete all columns (clear content)
				ws.delete_rows(1, 2000)  # Delete all rows (clear content)
				if len(Voltageresults)>0:
					rows = dataframe_to_rows(Voltageresults,index=True, header=True)
					for (r_idx, row) in enumerate(rows):
						for (c_idx, value) in enumerate(row):
							ws.cell(row=r_idx+1, column=c_idx+1, value=value)

				wb.save(excel_file_path)


				# Access the sheet and clear its content
				ws = wb['CST_RAW']
				ws.delete_cols(1, 2000)  # Delete all columns (clear content)
				ws.delete_rows(1, 2000)  # Delete all rows (clear content)
				if len(Thermalresults)>0:
					rows = dataframe_to_rows(Thermalresults,index=True, header=True)
					for (r_idx, row) in enumerate(rows):
						for (c_idx, value) in enumerate(row):
							ws.cell(row=r_idx+1, column=c_idx+1, value=value)
					try:
						if len(ShiftFactor_df_pvt)>0:
							ws = wb['CAT A GEFF']
							ws.delete_cols(1, 2000)  # Delete all columns (clear content)
							ws.delete_rows(1, 2000)  # Delete all rows (clear content)
							rows = dataframe_to_rows(ShiftFactor_df_pvt,index=True, header=True)
							for (r_idx, row) in enumerate(rows):
								for (c_idx, value) in enumerate(row):
									ws.cell(row=r_idx+1, column=c_idx+1, value=value)		
					except:
						pass				

				wb.save(excel_file_path)
				logger.info(f"Themal results are saved in {outputFile}.")
				try:
					logger.info(f"Pivot tables are being created.")
					ws = wb['Pivot1_RAW']
					ws.delete_cols(1, 2000)  # Delete all columns (clear content)
					ws.delete_rows(1, 2000)  # Delete all rows (clear content)    
					projects = Thermalresults.columns[Thermalresults.columns.str.contains('#')].tolist()
					Thermalresults = Thermalresults.loc[Thermalresults["POWERFLOW_PostProject"]>Thermalresults['EmergencyRating']]
					table1 = Thermalresults[['FROMNUMBER','TONUMBER','ID','Projects_in_Study']+projects].copy()
					table1 = table1.groupby(by = ['FROMNUMBER','TONUMBER','ID','Projects_in_Study']).max()
					table1.reset_index(inplace=True)
					for project in projects:
						table1.loc[table1[project] <= 3, project] = 0
						table1.loc[table1[project] > 3, project] = 1
						table1[project] = table1[project].replace(np.nan, 0)
					table1 = table1.loc[(table1[projects]!=0).any(axis=1)]
					
					rows = dataframe_to_rows(table1,index=True, header=True)
					for (r_idx, row) in enumerate(rows):
						for (c_idx, value) in enumerate(row):
							ws.cell(row=r_idx+1, column=c_idx+1, value=value)
					wb.save(excel_file_path)
					ws = wb['Pivot2_RAW']
					ws.delete_cols(1, 2000)  # Delete all columns (clear content)
					ws.delete_rows(1, 2000)  # Delete all rows (clear content)                                
					table2 = Thermalresults[['FROMNUMBER','TONUMBER','ID', 'Contingency','Projects_in_Study']+projects].copy()
					table2 = table2.groupby(by = ['FROMNUMBER','TONUMBER','ID','Contingency','Projects_in_Study']).max()
					
					table2.reset_index(inplace=True) 
					for project in projects:
						table2.loc[table2[project] <= 3, project] = 0
						table2[project] = table2[project].replace(np.nan, 0)
					table2 = table2.loc[(table2[projects]!=0).any(axis=1)]

					rows = dataframe_to_rows(table2,index=True, header=True)                                 
					for (r_idx, row) in enumerate(rows):
						for (c_idx, value) in enumerate(row):
							ws.cell(row=r_idx+1, column=c_idx+1, value=value)   
					wb.save(excel_file_path)  
					logger.info(f"Pivot tables are saved in {outputFile}.")           
				except:
					logger.warning(f"Issue with creating pivot tables.")                  
			else:
				logger.info(f"There are no violation to report.")

			

		print("Success")
	except:
		print("Failed")

		
	


	psspy.progress_output(1,'',[0,0])
	logger.close()





