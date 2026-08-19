import os
import sys
import re
import shutil
import logging
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

# ============================================================
# PSS/E SETUP
# ============================================================

PSSE_PATH = r"C:\Program Files\PTI\PSSE35\35.6\PSSPY311"
sys.path.append(PSSE_PATH)
os.environ["PATH"] += ";" + PSSE_PATH

import psse35
import psspy
import pssarrays

psspy.psseinit(50000)

_i = psspy.getdefaultint()
_f = psspy.getdefaultreal()
_s = psspy.getdefaultchar()

# ============================================================
# STUDY INPUTS"T:\Rshared\Delivery Shared\Python Files\Sina\GitRepo\cas-automation-tools\CPAT GUI\Functions\Code_test\Tool Testing\P3135.con.txt"
# ============================================================

BASE_CASE = Path(r"T:\Rshared\Delivery Shared\Python Files\Sina\GitRepo\cas-automation-tools\CPAT GUI\Functions\Code_test\Tool Testing\SP_LG_Post _Project_V0_con.sav")
CON_FILE  = Path(r"T:\Rshared\Delivery Shared\Python Files\Sina\GitRepo\cas-automation-tools\CPAT GUI\Functions\Code_test\Tool Testing\P3135.con")

WORK_DIR = Path(r"C:\Users\ssoleimanifard\Downloads\ShahroozTest")
WORK_DIR.mkdir(exist_ok=True)

RUN_ID = datetime.now().strftime("%Y%m%d_%H%M%S")
EXCEL_REPORT = WORK_DIR / f"N_1_1_ACCC_Professional_Report_{RUN_ID}.xlsx"

HANDLE_ISLANDS = True

THERMAL_LIMIT = 100.0
LOW_VOLTAGE_LIMIT = 0.90
HIGH_VOLTAGE_LIMIT = 1.10

STUDY_AREAS = [31,33,40,60]

SUB_FILE = WORK_DIR / "study.sub"
MON_FILE = WORK_DIR / "study.mon"
DFX_FILE = WORK_DIR / "study.dfx"

FNSL_FIRST_OUTAGE = [1, 0, 0, 1, 1, 0, 99, 0]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# ============================================================
# BASIC CHECKING
# ============================================================

def check(ierr, message):
    if ierr != 0:
        logging.warning(f"{message} failed. ierr = {ierr}")
    return ierr


# ============================================================
# READ CONTINGENCY FILE
# ============================================================

def read_con(file_path):
    contingencies = []

    with open(file_path, "r") as f:
        lines = f.readlines()

    i = 0
    while i < len(lines):
        line = lines[i].strip()

        if line.upper().startswith("CONTINGENCY"):
            match = re.search(r"'([^']+)'", line)
            name = match.group(1) if match else line.replace("CONTINGENCY", "").strip()

            actions = []
            i += 1

            while i < len(lines) and not lines[i].strip().upper().startswith("END"):
                action = lines[i].strip()
                if action:
                    actions.append(action)
                i += 1

            contingencies.append({
                "name": name,
                "actions": actions
            })

        i += 1

    return contingencies


# ============================================================
# WRITE SINGLE-CONTINGENCY .CON FILE
# ============================================================

def write_con_file(contingencies, output_file):
    with open(output_file, "w") as f:
        f.write("COM\n")
        f.write("COM Second outage contingency file\n")
        f.write("COM\n")

        for con in contingencies:
            f.write(f"CONTINGENCY '{con['name']}'\n")

            for action in con["actions"]:
                f.write(f" {action}\n")

            f.write("END\n\n")

        # Required final END for DFAX/ACCC file parsing
        f.write("END\n")


# ============================================================
# APPLY FIRST OUTAGE MANUALLY
# ============================================================

def apply_first_outage(actions):
    """
    Applies first-outage actions from the .con file.

    Supported:
      - OPEN BRANCH FROM BUS x TO BUS y CKT 'id'
      - OPEN LINE FROM BUS x TO BUS y CKT 'id'
      - DISCONNECT BUS x
      - REMOVE MACHINE bus 'id'
    """

    for action in actions:

        action_clean = action.strip()
        action_upper = action_clean.upper()

        # ====================================================
        # OPEN BRANCH / OPEN LINE
        # ====================================================
        if "OPEN BRANCH" in action_upper or "OPEN LINE" in action_upper:

            match = re.search(
                r"FROM BUS\s+'?(\d+)'?\s+TO BUS\s+'?(\d+)'?\s+CKT\s+'?([^'\s]+)'?",
                action_clean,
                re.IGNORECASE
            )

            if not match:
                logging.warning(
                    f"Could not parse branch outage action: {action_clean}"
                )
                continue

            from_bus = int(match.group(1))
            to_bus = int(match.group(2))
            ckt = match.group(3).strip()

            ierr = psspy.branch_chng_3(
                from_bus,
                to_bus,
                ckt,
                [0, _i, _i, _i, _i, _i],
                [_f] * 12,
                [_f] * 12,
                _s
            )

            check(
                ierr,
                f"Opening branch/line {from_bus}-{to_bus}-{ckt}"
            )

        # ====================================================
        # DISCONNECT BUS
        # ====================================================
        elif "DISCONNECT BUS" in action_upper:

            match = re.search(
                r"DISCONNECT BUS\s+'?(\d+)'?",
                action_clean,
                re.IGNORECASE
            )

            if not match:
                logging.warning(
                    f"Could not parse bus disconnect action: {action_clean}"
                )
                continue

            bus = int(match.group(1))

            ierr = psspy.dscn(bus)

            check(
                ierr,
                f"Disconnecting bus {bus}"
            )

        # ====================================================
        # REMOVE MACHINE
        # ====================================================
        elif "REMOVE MACHINE" in action_upper:

            match = re.search(
                r"REMOVE MACHINE\s+'?(\d+)'?\s+['\"]?([^'\"]+)['\"]?",
                action_clean,
                re.IGNORECASE
            )

            if not match:
                logging.warning(
                    f"Could not parse machine outage action: {action_clean}"
                )
                continue

            bus = int(match.group(1))
            mach_id = match.group(2).strip()

            ierr = psspy.machine_chng_2(
                bus,
                mach_id,
                [0, _i, _i, _i, _i, _i],
                [_f] * 17
            )

            check(
                ierr,
                f"Removing machine {bus}-{mach_id}"
            )

        # ====================================================
        # UNSUPPORTED
        # ====================================================
        else:

            logging.warning(
                f"Unsupported first-outage action: {action_clean}"
            )


# ============================================================
# CREATE SIMPLE SUB / MON FILES
# ============================================================

def create_sub_mon_files(study_areas=STUDY_AREAS):

    with open(SUB_FILE, "w") as f:
        f.write("COM\n")
        f.write("COM Study area subsystem\n")
        f.write("COM\n")
        f.write("SUBSYSTEM 'STUDY_AREA'\n")

        for area in study_areas:
            f.write(f" AREA {area}\n")

        f.write("END\n")
        f.write("END\n")

    with open(MON_FILE, "w") as f:
        f.write("COM\n")
        f.write("COM Monitor file\n")
        f.write("COM\n")
        f.write("MONITOR BRANCHES IN SUBSYSTEM 'STUDY_AREA'\n")
        f.write("MONITOR VOLTAGE RANGE SUBSYSTEM 'STUDY_AREA' 0.90 1.10\n")
        f.write("END\n")


# ============================================================
# GENERATE DFAX
# ============================================================

def generate_dfax(con_file, dfx_file):

    if Path(dfx_file).exists():
        try:
            Path(dfx_file).unlink()
        except Exception as e:
            logging.warning(f"Could not delete old DFX file: {dfx_file}. Error: {e}")
            return 999

    ierr = psspy.dfax_2(
        [1, 1, 1],
        str(SUB_FILE),
        str(MON_FILE),
        str(con_file),
        str(dfx_file)
    )

    check(ierr, "DFAX generation")

    if ierr != 0:
        return ierr

    if not Path(dfx_file).exists() or Path(dfx_file).stat().st_size == 0:
        logging.warning(f"DFAX did not create a valid file: {dfx_file}")
        return 998

    return 0


# ============================================================
# RUN ACCC
# ============================================================

def run_accc(acc_file, dfx_file):

    ierr = psspy.accc_with_dsp_3(
        0.5,
        [0, 0, 0, 1, 1, 1, 0, 0, 0, 0, 0],
        "",
        str(dfx_file),
        str(acc_file),
        "",
        "",
        ""
    )

    if ierr != 0:
        logging.warning(f"ACCC run failed. ierr = {ierr}")
    else:
        logging.info(f"ACCC completed successfully: {acc_file}")

    return ierr
# ============================================================
# BRANCH RATINGS / ACCC ELEMENT PARSING
# ============================================================

def normalize_ckt(ckt):
    """Normalize PSS/E circuit IDs for matching."""
    return str(ckt).strip().strip("'").strip('"')


def make_branch_key(from_bus, to_bus, ckt):
    return f"{int(from_bus)}-{int(to_bus)}-{normalize_ckt(ckt)}"


def build_branch_rating_lookup():
    """
    Builds a rating lookup from the base case.

    Keys are stored in both directions:
        FROM-TO-CKT
        TO-FROM-CKT

    The report uses RATE1 because your PSS/E loading basis is:
        Current expressed as MVA / RATE1
    """

    rating_lookup = {}

    ierr1, from_bus = psspy.abrnint(-1, 1, 1, 1, 1, ["FROMNUMBER"])
    ierr2, to_bus = psspy.abrnint(-1, 1, 1, 1, 1, ["TONUMBER"])
    ierr3, ckt = psspy.abrnchar(-1, 1, 1, 1, 1, ["ID"])

    # Try RATE1 first. If not available, fall back to RATEA.
    ierr4, rate = psspy.abrnreal(-1, 1, 1, 1, 1, ["RATE1"])
    rating_name = "RATE1"

    if ierr4 != 0:
        logging.warning("RATE1 lookup failed. Falling back to RATEA.")
        ierr4, rate = psspy.abrnreal(-1, 1, 1, 1, 1, ["RATEA"])
        rating_name = "RATEA"

    if ierr1 != 0 or ierr2 != 0 or ierr3 != 0 or ierr4 != 0:
        logging.warning("Could not build branch rating lookup.")
        return rating_lookup, rating_name

    fb = from_bus[0]
    tb = to_bus[0]
    ids = ckt[0]
    rates = rate[0]

    for f, t, c, r in zip(fb, tb, ids, rates):
        c = normalize_ckt(c)
        key1 = make_branch_key(f, t, c)
        key2 = make_branch_key(t, f, c)

        rating_lookup[key1] = float(r)
        rating_lookup[key2] = float(r)

    logging.info(f"Branch rating lookup built using {rating_name}. Number of keys = {len(rating_lookup)}")
    return rating_lookup, rating_name


def get_branch_key_from_accc_element(element):
    """
    Converts an ACCC monitored element label into:
        FROMBUS-TOBUS-CKT

    Expected ACCC branch label examples:
        266 EMPRESA7 138.00 674 CYPRES1 138.00 68
        136 E EDMON4 240.00 89 E EDMON7 138.00 T1

    This intentionally skips labels that do not look like normal two-terminal branches.
    """

    text = str(element).strip()
    parts = text.split()

    if len(parts) < 7:
        return None

    try:
        from_bus = int(parts[0])

        # The to-bus normally appears after the from-bus name and from-bus kV.
        # Find the next integer token after position 0.
        to_bus = None
        to_index = None
        for idx in range(1, len(parts)):
            if re.fullmatch(r"\d+", parts[idx]):
                to_bus = int(parts[idx])
                to_index = idx
                break

        if to_bus is None:
            return None

        # Circuit ID is normally the last token in the ACCC element label.
        ckt = normalize_ckt(parts[-1])

        # Skip obvious non-branch labels.
        if not ckt:
            return None

        return make_branch_key(from_bus, to_bus, ckt)

    except Exception:
        return None


# ============================================================
# EXCEL REPORTING
# ============================================================

def parse_acc_file(first_outage, acc_file, rating_lookup, rating_name):

    thermal_rows = []
    voltage_rows = []
    non_converged_rows = []
    all_case_rows = []
    diagnostic_rows = []

    try:
        summary = pssarrays.accc_summary(str(acc_file))

        contingencies = list(summary["colabel"])
        monitored_elements = list(summary["melement"])
        monitored_buses = list(summary["mvbuslabel"])

    except Exception as e:
        non_converged_rows.append({
            "First Outage": first_outage,
            "Second Outage": "",
            "Issue": f"Could not read ACC file: {e}",
            "ACC File": str(acc_file)
        })
        return thermal_rows, voltage_rows, non_converged_rows, all_case_rows, diagnostic_rows

    for second_outage in contingencies:

        try:
            sol = pssarrays.accc_solution(str(acc_file), second_outage)

            ierr = sol["ierr"]
            converged = sol["cnvflag"]
            cnvcond = sol["cnvcond"]
            island = sol["island"]

            loading_values = list(sol.get("ampflow", []))
            volts = list(sol.get("volts", []))

            all_case_rows.append({
                "First Outage": first_outage,
                "Second Outage": second_outage,
                "Converged": converged,
                "Convergence Condition": cnvcond,
                "Island": island,
                "Thermal Elements in ACCC": len(monitored_elements),
                "Thermal Values in Solution": len(loading_values),
                "Voltage Buses in ACCC": len(monitored_buses),
                "Voltage Values in Solution": len(volts),
                "ACC File": str(acc_file)
            })

            if ierr != 0 or not converged:
                non_converged_rows.append({
                    "First Outage": first_outage,
                    "Second Outage": second_outage,
                    "Issue": cnvcond,
                    "Island": island,
                    "ACC File": str(acc_file)
                })
                continue

            thermal_checked = 0
            thermal_missing_key = 0
            thermal_missing_rating = 0
            max_pct = 0.0
            max_element = ""
            max_key = ""
            max_loading = 0.0
            max_rating = 0.0

            # ------------------------------------------------
            # Thermal reporting
            # PSS/E option: Current expressed as MVA
            # Use ampflow / RATE1 * 100
            # ------------------------------------------------
            for element, loading_basis_value in zip(monitored_elements, loading_values):

                try:
                    thermal_checked += 1
                    loading_basis_value = abs(float(loading_basis_value))

                    branch_key = get_branch_key_from_accc_element(element)
                    if branch_key is None:
                        thermal_missing_key += 1
                        continue

                    rating = rating_lookup.get(branch_key)
                    if rating is None or rating <= 0:
                        thermal_missing_rating += 1
                        continue

                    loading_percent = loading_basis_value / rating * 100.0

                    if loading_percent > max_pct:
                        max_pct = loading_percent
                        max_element = element
                        max_key = branch_key
                        max_loading = loading_basis_value
                        max_rating = rating

                    if loading_percent >= THERMAL_LIMIT:
                        thermal_rows.append({
                            "First Outage": first_outage,
                            "Second Outage": second_outage,
                            "Violated Element": element,
                            "Branch Key": branch_key,
                            "Flow Basis": "Current expressed as MVA",
                            "Loading Basis Value": loading_basis_value,
                            "Rating Used": rating_name,
                            "Rating MVA": rating,
                            "Loading (%)": loading_percent,
                            "Limit (%)": THERMAL_LIMIT,
                            "Island": island,
                            "ACC File": str(acc_file)
                        })

                except Exception as e:
                    thermal_missing_key += 1
                    continue

            diagnostic_rows.append({
                "First Outage": first_outage,
                "Second Outage": second_outage,
                "Thermal Elements Checked": thermal_checked,
                "Missing/Unparsed Branch Key": thermal_missing_key,
                "Missing Rating": thermal_missing_rating,
                "Max Loading (%) Seen": max_pct,
                "Max Loading Element": max_element,
                "Max Loading Branch Key": max_key,
                "Max Loading Basis Value": max_loading,
                "Max Rating": max_rating,
                "Rating Used": rating_name,
                "ACC File": str(acc_file)
            })

            # ------------------------------------------------
            # Voltage reporting
            # ------------------------------------------------
            for bus, voltage in zip(monitored_buses, volts):
                try:
                    voltage_value = float(voltage)

                    if voltage_value <= LOW_VOLTAGE_LIMIT or voltage_value >= HIGH_VOLTAGE_LIMIT:
                        voltage_rows.append({
                            "First Outage": first_outage,
                            "Second Outage": second_outage,
                            "Bus": bus,
                            "Voltage (pu)": voltage_value,
                            "Low Limit": LOW_VOLTAGE_LIMIT,
                            "High Limit": HIGH_VOLTAGE_LIMIT,
                            "Island": island,
                            "ACC File": str(acc_file)
                        })

                except Exception:
                    continue

        except Exception as e:
            non_converged_rows.append({
                "First Outage": first_outage,
                "Second Outage": second_outage,
                "Issue": str(e),
                "ACC File": str(acc_file)
            })

    return thermal_rows, voltage_rows, non_converged_rows, all_case_rows, diagnostic_rows


def auto_format_excel(writer):
    workbook = writer.book

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        if ws.max_row < 1:
            continue

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            ws.column_dimensions[column_letter].width = min(max_length + 3, 70)

        ws.freeze_panes = "A2"


def export_professional_excel(acc_results):

    # Reload the base case before building ratings so ratings are not taken
    # from the final N-1 case left in memory.
    ierr = psspy.case(str(BASE_CASE))
    check(ierr, "Reloading base case before building rating lookup")

    rating_lookup, rating_name = build_branch_rating_lookup()

    all_thermal = []
    all_voltage = []
    all_non_converged = []
    all_cases = []
    all_diagnostics = []

    for item in acc_results:
        first_outage = item["First Outage"]
        acc_file = item["ACC File"]

        thermal, voltage, non_converged, cases, diagnostics = parse_acc_file(
            first_outage,
            acc_file,
            rating_lookup,
            rating_name
        )

        all_thermal.extend(thermal)
        all_voltage.extend(voltage)
        all_non_converged.extend(non_converged)
        all_cases.extend(cases)
        all_diagnostics.extend(diagnostics)

    df_thermal = pd.DataFrame(all_thermal)
    df_voltage = pd.DataFrame(all_voltage)
    df_non_converged = pd.DataFrame(all_non_converged)
    df_cases = pd.DataFrame(all_cases)
    df_files = pd.DataFrame(acc_results)
    df_diagnostics = pd.DataFrame(all_diagnostics)

    summary = pd.DataFrame([
        ["Total ACCC Files", len(df_files)],
        ["Total N-1-1 Cases Processed", len(df_cases)],
        ["Thermal Violations", len(df_thermal)],
        ["Voltage Violations", len(df_voltage)],
        ["Non-Converged / Error Cases", len(df_non_converged)],
        ["Rating Used for Thermal %", rating_name],
        ["Thermal Limit (%)", THERMAL_LIMIT],
        ["Low Voltage Limit (pu)", LOW_VOLTAGE_LIMIT],
        ["High Voltage Limit (pu)", HIGH_VOLTAGE_LIMIT]
    ], columns=["Item", "Value"])

    logging.info(f"Number of ACC files parsed: {len(acc_results)}")
    logging.info(f"Number of thermal violation rows: {len(df_thermal)}")
    logging.info(f"Number of voltage violation rows: {len(df_voltage)}")

    if not df_diagnostics.empty:
        max_seen = df_diagnostics["Max Loading (%) Seen"].max()
        logging.info(f"Maximum thermal loading seen in ACCC results: {max_seen}")

    with pd.ExcelWriter(EXCEL_REPORT, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        df_thermal.to_excel(writer, sheet_name="Thermal Violations", index=False)
        df_voltage.to_excel(writer, sheet_name="Voltage Violations", index=False)
        df_non_converged.to_excel(writer, sheet_name="Non-Converged Cases", index=False)
        df_cases.to_excel(writer, sheet_name="All Cases", index=False)
        df_diagnostics.to_excel(writer, sheet_name="Thermal Diagnostics", index=False)
        df_files.to_excel(writer, sheet_name="ACCC Files", index=False)

        auto_format_excel(writer)

    logging.info(f"Professional Excel report created: {EXCEL_REPORT}")


# ============================================================
# HANDLE ISLANDS
# ============================================================
def handle_islands(first_outage):
    """
    Checks for swingless islands.
    If found, disconnects each island using TREE,
    then returns True so the case can be solved and ACCC can continue.
    """

    ierr, buses = psspy.tree(1, 0)

    if ierr != 0:
        logging.warning(f"TREE initialization failed after first outage: {first_outage}. ierr = {ierr}")
        return False

    if buses == 0:
        logging.info(f"No island detected after first outage: {first_outage}")
        return True

    total_islanded_buses = 0

    while buses > 0:
        logging.warning(
            f"Swingless island detected after first outage: {first_outage}. "
            f"Island buses = {buses}. Disconnecting island."
        )

        total_islanded_buses += buses

        ierr, buses = psspy.tree(2, 1)

        if ierr != 0:
            logging.warning(f"TREE island disconnection failed after first outage: {first_outage}. ierr = {ierr}")
            return False

    logging.info(
        f"TREE completed after first outage: {first_outage}. "
        f"Total disconnected islanded buses = {total_islanded_buses}"
    )

    return True

# ============================================================
# MAIN N-1-1 ACCC STYLE STUDY
# ============================================================

def main():

    create_sub_mon_files(study_areas=STUDY_AREAS)

    contingencies = read_con(CON_FILE)
    logging.info(f"Total contingencies found: {len(contingencies)}")

    acc_results = []

    for idx, first_con in enumerate(contingencies):

        first_name = first_con["name"]
        safe_name = re.sub(r"[^A-Za-z0-9_]+", "_", first_name)

        logging.info(f"Running first outage {idx + 1}/{len(contingencies)}: {first_name}")

        # ----------------------------------------------------
        # Load base case
        # ----------------------------------------------------
        ierr = psspy.case(str(BASE_CASE))
        check(ierr, "Loading base case")

        if ierr != 0:
            logging.warning(f"Skipping {first_name} because base case could not be loaded.")
            continue

        # ----------------------------------------------------
        # Apply first outage ONCE
        # ----------------------------------------------------
        apply_first_outage(first_con["actions"])

        # ----------------------------------------------------
        # Handle islands immediately after first outage
        # ----------------------------------------------------
        case_ok = handle_islands(first_name)

        if not case_ok:
            logging.warning(f"Skipping first outage due to TREE/islanding issue before solve: {first_name}")
            continue

        # ----------------------------------------------------
        # Solve first-outage case
        # ----------------------------------------------------
        ierr = psspy.fnsl(FNSL_FIRST_OUTAGE)
        check(ierr, f"Solving first outage case: {first_name}")

        # ----------------------------------------------------
        # Check and handle islands again after solve
        # ----------------------------------------------------
        case_ok = handle_islands(first_name)

        if not case_ok:
            logging.warning(f"Skipping first outage due to TREE/islanding issue after solve: {first_name}")
            continue

        # ----------------------------------------------------
        # Re-solve after island handling
        # ----------------------------------------------------
        ierr = psspy.fnsl(FNSL_FIRST_OUTAGE)
        check(ierr, f"Re-solving first outage case after TREE check: {first_name}")

        if ierr != 0:
            logging.warning(f"Skipping first outage due to non-convergence: {first_name}")
            continue

        # ----------------------------------------------------
        # Save temporary N-1 case
        # ----------------------------------------------------
        temp_case = WORK_DIR / f"N1_{idx + 1}_{safe_name}.sav"
        ierr = psspy.save(str(temp_case))
        check(ierr, f"Saving temporary case for {first_name}")

        if ierr != 0:
            logging.warning(f"Skipping ACCC because N-1 case could not be saved: {first_name}")
            continue

        # ----------------------------------------------------
        # Create second-outage list
        # Exclude the first outage itself
        # ----------------------------------------------------
        second_conts = [
            c for c in contingencies
            if c["name"] != first_name
        ]

        second_con_file = WORK_DIR / f"second_outages_after_{idx + 1}_{safe_name}.con"
        write_con_file(second_conts, second_con_file)

        # ----------------------------------------------------
        # Optional debug: print generated second contingency file
        # Comment this block out when no longer needed
        # ----------------------------------------------------
        # with open(second_con_file, "r") as f:
        #     print("\n===== SECOND CON FILE =====")
        #     print(f.read())
        #     print("===========================\n")

        # ----------------------------------------------------
        # Generate DFAX for second outages
        # ----------------------------------------------------
        dfx_file = WORK_DIR / f"study_{idx + 1}_{safe_name}.dfx"

        ierr_dfx = generate_dfax(second_con_file, dfx_file)

        if ierr_dfx != 0:
            logging.warning(f"Skipping ACCC because DFAX failed for: {first_name}")
            continue

        acc_file = WORK_DIR / f"N1_1_after_{idx + 1}_{safe_name}.acc"

        if acc_file.exists():
            try:
                acc_file.unlink()
            except Exception as e:
                logging.warning(f"Could not delete old ACC file: {acc_file}. Error: {e}")
                continue

        ierr = run_accc(acc_file, dfx_file)

        if ierr == 0:
            logging.info(f"ACCC result saved: {acc_file}")
            acc_results.append({
                "First Outage": first_name,
                "ACC File": str(acc_file)
            })
        else:
            logging.warning(f"ACCC did not create a valid ACCC file: {acc_file}")

    export_professional_excel(acc_results)

    logging.info("N-1-1 ACCC-style study complete.")


if __name__ == "__main__":
    main()
